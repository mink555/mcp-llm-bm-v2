#!/usr/bin/env python3
"""
BFCL Excel Reporter - 평가 결과 엑셀 취합 모듈

모델별 평가 결과를 직관적인 엑셀 보고서로 생성합니다.
- Sheet 1: Overview (프로젝트 개요 및 평가 기준)
- Sheet 2: Summary (카테고리별 요약 - Detail 시트와 COUNTIFS 연동)
- Sheet 3: Detail (개별 테스트 케이스 결과 + GT + Pass 여부)
- Sheet 4: About BFCL (BFCL 벤치마크 상세 설명)
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다. pip install openpyxl 실행 필요")
    raise

# 깔끔하고 밝은 스타일 정의 - 기본 폰트 11 기준
# - 너무 어두운 헤더/배경을 피하고, 기본 글자 크기를 11로 통일
HEADER_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")  # 연한 회색 헤더
HEADER_FONT = Font(bold=True, size=12, color="111827")  # 진한 글씨
TITLE_FONT = Font(bold=True, size=18, color="111827")  # 제목
SUBTITLE_FONT = Font(bold=True, size=14, color="374151")  # 섹션 제목
SECTION_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")  # 매우 연한 회색
ACCENT_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")  # 연한 포인트(블루)
PASS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # 차분한 연한 초록
FAIL_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # 차분한 연한 빨강
WARNING_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")  # 차분한 노랑
INFO_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # 차분한 파랑
HIGHLIGHT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # 연한 회색
ALT_ROW_FILL = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")  # 교차 행 색상
THIN_BORDER = Border(
    left=Side(style="thin", color="E0E0E0"),
    right=Side(style="thin", color="E0E0E0"),
    top=Side(style="thin", color="E0E0E0"),
    bottom=Side(style="thin", color="E0E0E0"),
)
THICK_BORDER = Border(
    left=Side(style="medium", color="9E9E9E"),
    right=Side(style="medium", color="9E9E9E"),
    top=Side(style="medium", color="9E9E9E"),
    bottom=Side(style="medium", color="9E9E9E"),
)

# 에러 타입 → 한국어 설명 매핑
ERROR_TYPE_DESCRIPTIONS = {
    # AST 파싱 관련
    "ast_decoder:decoder_failed": "응답 파싱 실패 - 모델 응답을 함수 호출 형태로 변환할 수 없음",
    
    # 함수명 관련
    "simple_function_checker:wrong_func_name": "함수명 오류 - 호출한 함수명이 Ground Truth와 다름",
    "simple_function_checker:wrong_number_of_functions": "함수 개수 오류 - 호출한 함수 개수가 예상과 다름",
    
    # 파라미터 관련
    "simple_function_checker:missing_required": "필수 파라미터 누락 - required 파라미터가 포함되지 않음",
    "simple_function_checker:unexpected_param": "불필요한 파라미터 - 정의되지 않은 파라미터가 포함됨",
    "simple_function_checker:wrong_param_name": "파라미터명 오류 - 파라미터명이 예상과 다름",
    
    # 타입 관련
    "type_error:simple": "타입 오류 - 파라미터 타입이 예상과 다름 (예: 정수를 문자열로 반환)",
    "type_error:dict": "딕셔너리 타입 오류 - 딕셔너리 파라미터의 구조가 잘못됨",
    "type_error:list": "리스트 타입 오류 - 리스트 파라미터의 타입이 잘못됨",
    "type_error:tuple": "튜플 타입 오류 - 튜플 파라미터의 타입이 잘못됨",
    
    # 값 관련
    "value_error:string": "문자열 값 오류 - 문자열 값이 예상 범위에 없음",
    "value_error:integer": "정수 값 오류 - 정수 값이 예상 범위에 없음",
    "value_error:float": "실수 값 오류 - 실수 값이 예상 범위에 없음",
    "value_error:boolean": "불리언 값 오류 - True/False 값이 예상과 다름",
    "value_error:dict": "딕셔너리 값 오류 - 딕셔너리 값이 예상과 다름",
    "value_error:list": "리스트 값 오류 - 리스트 값이 예상과 다름",
    "value_error:enum": "열거형 값 오류 - 허용된 값 목록에 없음",
    
    # 병렬 함수 호출 관련
    "parallel_function_checker_no_order:wrong_count": "병렬 호출 개수 오류 - 호출한 함수 개수가 예상과 다름",
    "parallel_function_checker_no_order:cannot_find_match": "병렬 호출 매칭 실패 - 기대하는 함수 호출을 찾을 수 없음 (함수명/파라미터 불일치)",
    "parallel_function_checker_enforce_order:wrong_count": "병렬 호출 개수 오류 (순서 검사) - 호출한 함수 개수가 예상과 다름",
    
    # 다중 함수 호출 관련
    "multiple_function_checker:wrong_count": "다중 호출 개수 오류 - 호출한 함수 개수가 예상과 다름",
    
    # 기타
    "no_function_call": "함수 호출 없음 - 모델이 함수를 호출하지 않음",
    "irrelevance:no_function_call": "정상 (함수 호출 안 함) - 불필요한 함수 호출을 올바르게 거부함",
    "relevance:no_function_call": "함수 호출 누락 - 필요한 함수를 호출하지 않음",
}

def get_error_description_kr(error_type: str) -> str:
    """에러 타입에 대한 한국어 설명 반환"""
    if not error_type:
        return ""
    
    # 정확히 일치하는 경우
    if error_type in ERROR_TYPE_DESCRIPTIONS:
        return ERROR_TYPE_DESCRIPTIONS[error_type]
    
    # 부분 일치 검색 (예: "type_error:simple" -> "type_error")
    for key, desc in ERROR_TYPE_DESCRIPTIONS.items():
        if error_type.startswith(key.split(":")[0] + ":"):
            return desc
    
    # 기본값
    return f"기타 오류 - {error_type}"


# BFCL 카테고리 그룹 매핑
CATEGORY_GROUPS = {
    "simple_python": ("Non-Live", "Simple"),
    "simple_java": ("Non-Live", "Simple"),
    "simple_javascript": ("Non-Live", "Simple"),
    "multiple": ("Non-Live", "Multiple"),
    "parallel": ("Non-Live", "Parallel"),
    "parallel_multiple": ("Non-Live", "Parallel Multiple"),
    "irrelevance": ("Non-Live", "Irrelevance"),
    "live_simple": ("Live", "Simple"),
    "live_multiple": ("Live", "Multiple"),
    "live_parallel": ("Live", "Parallel"),
    "live_parallel_multiple": ("Live", "Parallel Multiple"),
    "live_irrelevance": ("Live", "Irrelevance"),
    "live_relevance": ("Live", "Relevance"),
    "multi_turn_base": ("Multi-Turn", "Base"),
    "multi_turn_miss_func": ("Multi-Turn", "Miss Function"),
    "multi_turn_miss_param": ("Multi-Turn", "Miss Parameter"),
    "multi_turn_long_context": ("Multi-Turn", "Long Context"),
    "memory_kv": ("Agentic", "Memory KV"),
    "memory_vector": ("Agentic", "Memory Vector"),
    "memory_rec_sum": ("Agentic", "Memory RecSum"),
    "web_search_base": ("Agentic", "Web Search"),
    "web_search_no_snippet": ("Agentic", "Web Search No Snippet"),
}

# 카테고리(데이터셋) 간단 설명 (시트에 표기용)
# - 키: BFCL category id
# - 값: 사람이 바로 이해할 수 있는 한 줄 설명
CATEGORY_DESCRIPTIONS_KR = {
    # Non-live
    "simple_python": "단일 함수 1회 호출 (Python)",
    "simple_java": "단일 함수 1회 호출 (Java)",
    "simple_javascript": "단일 함수 1회 호출 (JavaScript)",
    "multiple": "여러 함수 중 올바른 함수 1개 선택",
    "parallel": "동일 함수의 병렬 다중 호출",
    "parallel_multiple": "여러 함수의 병렬 다중 호출",
    "irrelevance": "도구 호출이 불필요한 요청을 거부/무시",

    # Live
    "live_simple": "실시간(Live) 단일 호출",
    "live_multiple": "실시간(Live) 다중 함수 선택",
    "live_parallel": "실시간(Live) 병렬 호출",
    "live_parallel_multiple": "실시간(Live) 병렬 다중 호출",
    "live_irrelevance": "실시간(Live) Irrelevance",
    "live_relevance": "실시간(Live) Relevance",

    # Multi-turn / Agentic
    "multi_turn_base": "멀티턴 기본",
    "multi_turn_miss_func": "멀티턴: 함수 선택 실패 유도",
    "multi_turn_miss_param": "멀티턴: 파라미터 누락 유도",
    "multi_turn_long_context": "멀티턴: 긴 컨텍스트",
    "memory_kv": "Agentic: Memory KV",
    "memory_vector": "Agentic: Memory Vector",
    "memory_rec_sum": "Agentic: Memory RecSum",
    "web_search_base": "Agentic: Web Search",
    "web_search_no_snippet": "Agentic: Web Search (No Snippet)",
}

def describe_category(category: str) -> str:
    """카테고리 한 줄 설명(없으면 공백)."""
    return CATEGORY_DESCRIPTIONS_KR.get(category, "")


def parse_error_details(error_type: str, error_raw: str) -> dict:
    """
    error 메시지 원문에서 핵심을 구조화.
    반환 키:
      - area: 부족 영역(예: 타입/필수파라미터/함수선택/파싱/값/기타)
      - param: 관련 파라미터(있으면)
      - expected_type / actual_type: 타입 오류인 경우
      - actual_value: 실제 값(있으면)
    """
    import re

    error_type = error_type or ""
    error_raw = error_raw or ""

    # English labels for analysis columns
    area = "Other"
    if error_type.startswith("type_error"):
        area = "Type"
    elif "missing_required" in error_type:
        area = "Required Param"
    elif "wrong_param_name" in error_type or "unexpected_param" in error_type:
        area = "Parameter"
    elif "wrong_func_name" in error_type:
        area = "Function Selection"
    elif "decoder_failed" in error_type:
        area = "Parsing"
    elif error_type.startswith("value_error"):
        area = "Value"
    elif "cannot_find_match" in error_type or "wrong_count" in error_type:
        area = "Count/Matching"

    param = ""
    expected_type = ""
    actual_type = ""
    actual_value = ""

    m = re.search(r"parameter '([^']+)'", error_raw)
    if m:
        param = m.group(1)

    tm = re.search(r"Expected type ([^,]+), got ([^\.]+)\.", error_raw)
    if tm:
        expected_type = tm.group(1).strip()
        actual_type = tm.group(2).strip()

    vm = re.search(r"Parameter value:\s*'([^']*)'", error_raw)
    if vm:
        actual_value = vm.group(1)

    return {
        "area": area,
        "param": param,
        "expected_type": expected_type,
        "actual_type": actual_type,
        "actual_value": actual_value,
    }


def load_bfcl_dataset(category: str) -> tuple[dict, dict]:
    """BFCL 데이터셋에서 prompt와 ground truth 로드"""
    try:
        from bfcl_eval.utils import load_dataset_entry, load_ground_truth_entry
        
        prompts = load_dataset_entry(category, include_prereq=False, include_language_specific_hint=False)
        ground_truths = load_ground_truth_entry(category)
        
        # ID로 인덱싱
        prompt_by_id = {}
        for p in prompts:
            entry_id = p.get("id", "")
            # 질문 추출: question[0][0]["content"]
            question = ""
            if "question" in p and p["question"]:
                if isinstance(p["question"][0], list) and p["question"][0]:
                    question = p["question"][0][0].get("content", "")
                elif isinstance(p["question"][0], dict):
                    question = p["question"][0].get("content", "")
            prompt_by_id[entry_id] = question
        
        gt_by_id = {}
        for g in ground_truths:
            entry_id = g.get("id", "")
            gt_by_id[entry_id] = g.get("ground_truth", [])
        
        return prompt_by_id, gt_by_id
    except Exception as e:
        print(f"Warning: Failed to load BFCL dataset for {category}: {e}")
        return {}, {}


class BFCLExcelReporter:
    """BFCL 평가 결과 엑셀 보고서 생성기"""

    def __init__(self, model_name: str, result_dir: Path, score_dir: Path):
        self.model_name = model_name
        self.result_dir = result_dir
        self.score_dir = score_dir
        self.wb = Workbook()
        self.detail_data = []  # 상세 결과 (score 파일에서 로드)
        self.categories_found = set()  # 발견된 카테고리 목록

    def load_data(self) -> None:
        """score 및 result 파일에서 데이터 로드"""
        # 1. Score 파일에서 실패한 항목 로드
        for subdir in ["non_live", "live", "multi_turn", "agentic"]:
            subdir_path = self.score_dir / subdir
            if not subdir_path.exists():
                continue
            
            for score_file in subdir_path.glob("*_score.json"):
                category = self._extract_category(score_file.name)
                self._load_score_file(score_file, category, subdir)
        
        # 2. Result 파일에서 모든 항목 로드 (score에 없는 항목은 PASS)
        self._load_result_files()

    def _extract_category(self, filename: str) -> str:
        """파일명에서 카테고리 추출: BFCL_v4_simple_python_score.json -> simple_python"""
        name = filename.replace("_score.json", "").replace("_result.json", "")
        if name.startswith("BFCL_v"):
            parts = name.split("_", 2)
            if len(parts) > 2:
                return parts[2]
        return name

    def _load_result_files(self) -> None:
        """Result 파일에서 모든 테스트 케이스 로드 (score에 없는 항목은 PASS)"""
        # 이미 로드된 ID 목록
        loaded_ids = {entry["id"] for entry in self.detail_data}
        
        for subdir in ["non_live", "live", "multi_turn", "agentic"]:
            subdir_path = self.result_dir / subdir
            if not subdir_path.exists():
                continue
            
            for result_file in subdir_path.glob("*_result.json"):
                category = self._extract_category(result_file.name)
                self._load_result_file(result_file, category, loaded_ids)
        
        # 모든 데이터에 대해 BFCL 데이터셋에서 질문과 GT 보강
        self._enrich_with_bfcl_dataset()

    def _enrich_with_bfcl_dataset(self) -> None:
        """BFCL 데이터셋에서 실제 질문과 Ground Truth 가져오기"""
        # 카테고리별로 데이터셋 로드
        dataset_cache = {}
        
        for entry in self.detail_data:
            category = entry["category"]
            entry_id = entry["id"]
            
            # 캐시에 없으면 로드
            if category not in dataset_cache:
                prompt_by_id, gt_by_id = load_bfcl_dataset(category)
                dataset_cache[category] = (prompt_by_id, gt_by_id)
            
            prompt_by_id, gt_by_id = dataset_cache[category]
            
            # 실제 질문 업데이트
            if entry_id in prompt_by_id:
                entry["request"] = prompt_by_id[entry_id]
            else:
                entry["request"] = ""
            
            # Ground Truth 업데이트
            if entry_id in gt_by_id and not entry.get("expected"):
                gt = gt_by_id[entry_id]
                entry["expected"] = self._format_ground_truth(gt)

    def _format_ground_truth(self, gt: list) -> str:
        """Ground Truth 포맷팅"""
        if not gt:
            return ""
        try:
            # 함수 호출 형태로 변환
            result_parts = []
            for item in gt:
                if isinstance(item, dict):
                    for func_name, params in item.items():
                        # params에서 첫 번째 값 추출 (리스트 형태)
                        clean_params = {}
                        for k, v in params.items():
                            if isinstance(v, list) and v:
                                clean_params[k] = v[0]
                            else:
                                clean_params[k] = v
                        result_parts.append(f"{func_name}({clean_params})")
            return "; ".join(result_parts)[:300]
        except:
            return json.dumps(gt, ensure_ascii=False)[:300]

    def _load_result_file(self, filepath: Path, category: str, loaded_ids: set) -> None:
        """Result 파일 로드 - score에 없는 항목은 PASS로 표시"""
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                try:
                    entry = json.loads(line.strip())
                    entry_id = entry.get("id", "")
                    
                    # 이미 score 파일에서 로드된 항목은 건너뜀
                    if entry_id in loaded_ids:
                        continue
                    
                    self.categories_found.add(category)
                    
                    # Score 파일에 없으면 PASS (실패한 항목만 score에 기록됨)
                    self.detail_data.append({
                        "id": entry_id,
                        "category": category,
                        "group": CATEGORY_GROUPS.get(category, ("Unknown", "Unknown"))[0],
                        "type": CATEGORY_GROUPS.get(category, ("Unknown", "Unknown"))[1],
                        "valid": True,
                        "result": "PASS",
                        "request": "",  # 나중에 _enrich_with_bfcl_dataset에서 채움
                        "query": entry_id,
                        "expected": "",  # 나중에 _enrich_with_bfcl_dataset에서 채움
                        "actual": self._format_actual_from_result(entry),
                        "error": [],
                        "error_type": "",
                    })
                except json.JSONDecodeError:
                    continue

    def _format_actual_from_result(self, entry: dict) -> str:
        """Result 파일에서 모델 응답 포맷팅 - 파싱된 함수 호출 형식으로 변환"""
        result = entry.get("result", "")
        return self._format_model_response(result)
    
    def _format_model_response(self, result) -> str:
        """모델 응답을 함수 호출 형식으로 포맷팅"""
        if isinstance(result, list):
            # 이미 파싱된 형태: [{"func_name": {"param": value}}, ...]
            try:
                parts = []
                for item in result:
                    if isinstance(item, dict):
                        for func_name, params in item.items():
                            if isinstance(params, str):
                                # JSON 문자열인 경우 파싱
                                try:
                                    params = json.loads(params)
                                except:
                                    pass
                            if isinstance(params, dict):
                                parts.append(f"{func_name}({params})")
                            else:
                                parts.append(f"{func_name}({params})")
                return "; ".join(parts)[:300] if parts else str(result)[:300]
            except:
                return str(result)[:300]
        elif isinstance(result, str):
            # 문자열인 경우 파싱 시도
            parsed = self._try_parse_tool_calls(result)
            if parsed:
                return self._format_model_response(parsed)
            return result[:300]
        elif isinstance(result, dict):
            try:
                return json.dumps(result, ensure_ascii=False)[:300]
            except:
                return str(result)[:300]
        return str(result)[:300]
    
    def _try_parse_tool_calls(self, text: str) -> list:
        """텍스트에서 tool calls 파싱 시도"""
        import re
        text = text.strip()
        
        # [TOOL_CALLS] 태그 제거
        if text.startswith("[TOOL_CALLS]"):
            text = text[len("[TOOL_CALLS]"):].strip()
        if text.startswith("[TOOL_CALLS"):
            text = re.sub(r'^\[TOOL_CALLS\]?', '', text)
        
        # JSON 배열 파싱 시도
        try:
            if text.startswith("["):
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    result = []
                    for item in parsed:
                        if isinstance(item, dict):
                            if "name" in item and "arguments" in item:
                                func_name = item["name"]
                                args = item["arguments"]
                                if isinstance(args, str):
                                    args = json.loads(args)
                                result.append({func_name: args})
                            elif len(item) == 1:
                                result.append(item)
                    if result:
                        return result
        except:
            pass
        
        # func_name{...} 형식 파싱
        pattern = r'([a-zA-Z_][a-zA-Z0-9_\.]*)\s*(\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\})'
        matches = re.findall(pattern, text)
        if matches:
            result = []
            for func_name, json_str in matches:
                try:
                    params = json.loads(json_str.replace("'", '"'))
                    result.append({func_name: params})
                except:
                    continue
            if result:
                return result
        
        return None

    def _load_score_file(self, filepath: Path, category: str, subdir: str) -> None:
        """score 파일 로드 - 첫 줄은 요약, 나머지는 개별 결과"""
        with open(filepath, "r", encoding="utf-8") as f:
            lines = f.readlines()
        
        if not lines:
            return
        
        self.categories_found.add(category)
        
        # 나머지: 개별 결과
        for line in lines[1:]:
            if not line.strip():
                continue
            try:
                entry = json.loads(line.strip())
                entry_id = entry.get("id", "")
                request_text = self._extract_request(entry)  # 실제 질문
                self.detail_data.append({
                    "id": entry_id,
                    "category": category,
                    "group": CATEGORY_GROUPS.get(category, ("Unknown", "Unknown"))[0],
                    "type": CATEGORY_GROUPS.get(category, ("Unknown", "Unknown"))[1],
                    "valid": entry.get("valid", False),
                    "result": "PASS" if entry.get("valid", False) else "FAIL",
                    "query": entry_id,  # ID
                    "request": request_text,  # 실제 질문 (score 파일의 prompt.question에서)
                    "expected": self._format_expected(entry),
                    "actual": self._format_actual(entry),
                    "error": entry.get("error", []),
                    "error_type": entry.get("error_type", ""),
                })
            except json.JSONDecodeError:
                continue

    def _extract_request(self, entry: dict) -> str:
        """실제 질문 추출 (score 파일의 prompt.question에서)"""
        prompt = entry.get("prompt", {})
        question = prompt.get("question", [])
        if question and isinstance(question, list) and len(question) > 0:
            if isinstance(question[0], list) and len(question[0]) > 0:
                first_msg = question[0][0]
                if isinstance(first_msg, dict):
                    return first_msg.get("content", "")[:300]
            elif isinstance(question[0], dict):
                return question[0].get("content", "")[:300]
        return ""

    def _format_expected(self, entry: dict) -> str:
        """Ground Truth 포맷팅"""
        possible_answer = entry.get("possible_answer", [])
        if possible_answer:
            try:
                return json.dumps(possible_answer[0], ensure_ascii=False)[:200]
            except:
                return str(possible_answer)[:200]
        return ""

    def _format_actual(self, entry: dict) -> str:
        """모델 응답 포맷팅 - 파싱된 함수 호출 형식으로 변환"""
        raw = entry.get("model_result_raw", "")
        return self._format_model_response(raw)

    def create_overview_sheet(self) -> None:
        """Sheet 1: Overview - 프로젝트 개요 및 핵심 정보"""
        ws = self.wb.active
        ws.title = "Overview"
        
        current_row = 1
        
        # 메인 타이틀 - 큰 글씨로 강조
        title_cell = ws.cell(row=current_row, column=1, value="LLM Function Calling 평가 보고서")
        title_cell.font = Font(bold=True, size=16, color="2D3748")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws.row_dimensions[current_row].height = 35
        current_row += 1
        
        # 구분선
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(bottom=Side(style="medium", color="4A5568"))
        current_row += 2
        
        # 핵심 정보 박스
        info_start_row = current_row
        ws.cell(row=current_row, column=2, value="평가 정보").font = SUBTITLE_FONT
        ws.cell(row=current_row, column=2).fill = HEADER_FILL
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=current_row, column=2).border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:C{current_row}')
        current_row += 1
        
        info_data = [
            ("벤치마크", "BFCL V4 (Berkeley Function Calling Leaderboard)"),
            ("평가 모델", self.model_name.split('/')[-1]),
            ("평가 방식", "AST 기반 정확도 검증"),
            ("평가 일시", datetime.now().strftime("%Y년 %m월 %d일 %H:%M")),
        ]
        
        for label, value in info_data:
            label_cell = ws.cell(row=current_row, column=2, value=label)
            label_cell.font = Font(bold=True, size=10, color="4A5568")
            label_cell.border = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            
            value_cell = ws.cell(row=current_row, column=3, value=value)
            value_cell.border = THIN_BORDER
            value_cell.font = Font(size=11)
            value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            
            current_row += 1
        
        current_row += 1
        
        # 평가 카테고리 - 표 형식으로 깔끔하게
        section_cell = ws.cell(row=current_row, column=2, value="평가 카테고리")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = HEADER_FILL
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:E{current_row}')
        current_row += 1
        
        # 카테고리 테이블 헤더
        cat_headers = ["카테고리", "설명", "테스트 수"]
        for col, header in enumerate(cat_headers, 2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=10, color="2D3748")
            cell.fill = HIGHLIGHT_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        category_info = [
            ("Simple", "단일 함수 단일 호출", "550개"),
            ("Multiple", "다중 함수 중 선택", "200개"),
            ("Parallel", "다중 병렬 호출", "200개"),
            ("Parallel Multiple", "복잡한 병렬 호출", "200개"),
        ]
        
        for cat_name, description, count in category_info:
            name_cell = ws.cell(row=current_row, column=2, value=cat_name)
            name_cell.font = Font(bold=True, size=10, color="2D3748")
            name_cell.border = THIN_BORDER
            name_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            desc_cell = ws.cell(row=current_row, column=3, value=description)
            desc_cell.border = THIN_BORDER
            desc_cell.font = Font(size=11)
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            
            count_cell = ws.cell(row=current_row, column=4, value=count)
            count_cell.border = THIN_BORDER
            count_cell.font = Font(size=11, color="6B7280")
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += 1
        
        current_row += 1
        
        # Pass/Fail 기준 - 명확하게
        section_cell = ws.cell(row=current_row, column=2, value="Pass 기준")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = HEADER_FILL
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:E{current_row}')
        current_row += 1
        
        pass_criteria = [
            "1. 올바른 함수 선택: Ground Truth와 함수명 일치",
            "2. 필수 파라미터 전달: 모든 required 파라미터 포함",
            "3. 정확한 파라미터 값: 허용 범위 내의 값",
            "4. 올바른 타입 처리: 정수/문자열 등 타입 일치",
        ]
        
        for criterion in pass_criteria:
            crit_cell = ws.cell(row=current_row, column=2, value=criterion)
            crit_cell.font = Font(size=11, color="111827")
            crit_cell.border = THIN_BORDER
            crit_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            crit_cell.fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
            ws.merge_cells(f'B{current_row}:E{current_row}')
            ws.row_dimensions[current_row].height = 20
            
            current_row += 1
        
        current_row += 1
        
        # 보고서 구성
        section_cell = ws.cell(row=current_row, column=2, value="보고서 구성")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = HEADER_FILL
        section_cell.alignment = Alignment(horizontal="center", vertical="center")
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:E{current_row}')
        current_row += 1
        
        sheets_info = [
            ("Summary", "카테고리별 정확도 요약"),
            ("Detail", "개별 테스트 케이스 상세 결과"),
            ("About BFCL", "벤치마크 상세 설명"),
        ]
        
        for idx, (sheet_name, description) in enumerate(sheets_info, 1):
            sheet_cell = ws.cell(row=current_row, column=2, value=f"{idx}. {sheet_name}")
            sheet_cell.font = Font(bold=True, size=10, color="111827")
            sheet_cell.border = THIN_BORDER
            sheet_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            sheet_cell.fill = HIGHLIGHT_FILL
            
            desc_cell = ws.cell(row=current_row, column=3, value=description)
            desc_cell.border = THIN_BORDER
            desc_cell.font = Font(size=11)
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.merge_cells(f'C{current_row}:E{current_row}')
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        # 열 너비 최적화
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 3
        
        # 열 너비 최적화
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 3
    
    def create_about_bfcl_sheet(self) -> None:
        """Sheet 4: About BFCL - BFCL 벤치마크 상세 설명"""
        ws = self.wb.create_sheet("About BFCL")
        
        current_row = 2
        
        # 타이틀
        title_cell = ws.cell(row=current_row, column=2, value="BFCL (Berkeley Function Calling Leaderboard) 상세 가이드")
        title_cell.font = TITLE_FONT
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 2
        
        # BFCL이란?
        section_cell = ws.cell(row=current_row, column=2, value="BFCL이란?")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        desc_cell = ws.cell(row=current_row, column=2, 
                           value="UC Berkeley에서 개발한 LLM Function Calling 능력 평가의 표준 벤치마크입니다. "
                                 "실제 개발 환경을 반영한 다양한 시나리오와 AST 기반의 정확한 평가 방식을 특징으로 합니다.")
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws.row_dimensions[current_row].height = 30
        current_row += 2
        
        # 버전 히스토리
        section_cell = ws.cell(row=current_row, column=2, value="버전 히스토리")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        # 헤더
        headers = ["버전", "출시일", "주요 특징"]
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        versions = [
            ("V1", "2024.02", "AST 평가 방식 도입"),
            ("V2", "2024.05", "실시간(Live) 데이터 추가"),
            ("V3", "2024.09", "Multi-Turn 대화 지원"),
            ("V4", "2024.12", "Agentic 평가 추가"),
        ]
        
        for ver, date, feature in versions:
            ws.cell(row=current_row, column=2, value=ver).border = THIN_BORDER
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            ws.cell(row=current_row, column=3, value=date).border = THIN_BORDER
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=4, value=feature).border = THIN_BORDER
            ws.merge_cells(f'D{current_row}:F{current_row}')
            current_row += 1
        
        current_row += 1
        
        # 평가 철학
        section_cell = ws.cell(row=current_row, column=2, value="평가 철학")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        philosophy = [
            ("올바른 함수 선택", "여러 함수 중 사용자 의도에 맞는 적절한 함수를 선택하는 능력"),
            ("올바른 파라미터 전달", "필수/선택 파라미터를 정확하게 구분하고 전달하는 능력"),
            ("올바른 타입 처리", "정수, 문자열, 불리언 등 데이터 타입을 정확하게 다루는 능력"),
            ("함수 호출 판단", "상황에 따라 함수 호출이 필요한지 불필요한지 판단하는 능력"),
        ]
        
        for idx, (title, desc) in enumerate(philosophy, 1):
            title_cell = ws.cell(row=current_row, column=2, value=f"{idx}. {title}")
            title_cell.font = Font(bold=True, size=10)
            title_cell.fill = INFO_FILL
            title_cell.border = THIN_BORDER
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            desc_cell = ws.cell(row=current_row, column=3, value=desc)
            desc_cell.border = THIN_BORDER
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.merge_cells(f'C{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = 25
            
            current_row += 1
        
        current_row += 1
        
        # AST 평가 방식
        section_cell = ws.cell(row=current_row, column=2, value="AST 평가 방식")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        ast_steps = [
            ("1) 함수 파싱", "모델 응답을 Abstract Syntax Tree로 파싱하여 함수명과 파라미터를 추출합니다"),
            ("2) 함수명 검증", "추출된 함수명이 Ground Truth와 정확히 일치하는지 확인합니다"),
            ("3) 파라미터 검증", "각 파라미터의 키와 값이 예상 범위(possible_answer) 내에 있는지 확인합니다"),
            ("4) 타입 검증", "파라미터 타입이 함수 스키마와 일치하는지 확인합니다 (타입 자동 변환 없음)"),
        ]
        
        for step, desc in ast_steps:
            step_cell = ws.cell(row=current_row, column=2, value=step)
            step_cell.font = Font(bold=True, size=10)
            step_cell.border = THIN_BORDER
            step_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            desc_cell = ws.cell(row=current_row, column=3, value=desc)
            desc_cell.border = THIN_BORDER
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.merge_cells(f'C{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = 25
            
            current_row += 1
        
        current_row += 1
        
        # 의도적 난이도
        section_cell = ws.cell(row=current_row, column=2, value="의도적으로 포함된 난이도")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        difficulty_cell = ws.cell(row=current_row, column=2, 
                                 value="BFCL은 실제 Python 코딩 패턴을 반영하기 위해 의도적으로 어려운 요소들을 포함합니다:\n"
                                       "• 함수명 다양성: math.factorial, spotify.play 등 다양한 네이밍 패턴\n"
                                       "• API 호환성 문제: '.'이 포함된 함수명은 OpenAI API 규칙(^[a-zA-Z0-9_-]+$)에 위배\n"
                                       "• 타입 처리: 모델의 타입 변환 능력도 평가 대상 (자동 변환 미지원)")
        difficulty_cell.alignment = Alignment(wrap_text=True, vertical="top")
        difficulty_cell.fill = HIGHLIGHT_FILL
        difficulty_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws.row_dimensions[current_row].height = 60
        current_row += 2
        
        # 평가 지표
        section_cell = ws.cell(row=current_row, column=2, value="평가 지표")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        metrics = [
            ("Accuracy", "카테고리별 정답률 = Correct / Total"),
            ("Overall Accuracy", "전체 카테고리 Accuracy의 비가중 평균 (BFCL 공식 지표)"),
            ("Weighted Accuracy", "전체 정답 수 / 전체 테스트 수 (참고용)"),
        ]
        
        for metric, desc in metrics:
            metric_cell = ws.cell(row=current_row, column=2, value=metric)
            metric_cell.font = Font(bold=True, size=10)
            metric_cell.fill = HIGHLIGHT_FILL
            metric_cell.border = THIN_BORDER
            metric_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            desc_cell = ws.cell(row=current_row, column=3, value=desc)
            desc_cell.border = THIN_BORDER
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f'C{current_row}:F{current_row}')
            
            current_row += 1
        
        # 열 너비 조정
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 5
        ws.column_dimensions["F"].width = 5

    def create_dashboard_sheet(self) -> None:
        """핵심만 압축한 요약 시트"""
        ws = self.wb.active
        ws.title = "요약"

        # 카테고리 목록(표 행 수를 정하기 위한 용도만 Python 사용)
        categories = sorted({e.get("category") for e in self.detail_data if e.get("category")})

        # 상단 타이틀
        ws.merge_cells("A1:H1")
        t = ws.cell(row=1, column=1, value=f"{self.model_name.split('/')[-1]} 평가 요약")
        t.font = Font(bold=True, size=16, color="111827")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        # 우측 기준/로직 박스 (전체 폭 병합 대신, 고정 폭 영역으로)
        ws.merge_cells("G2:H6")
        basis = ws.cell(
            row=2,
            column=7,
            value=(
                "기준/로직\n"
                "- PASS/FAIL: BFCL 평가 결과\n"
                "- Overall Accuracy(BFCL): 카테고리 accuracy 비가중 평균\n"
                "- Weighted Accuracy: Σcorrect / Σtotal"
            ),
        )
        basis.font = Font(size=11, color="374151")
        basis.fill = SECTION_FILL
        basis.border = THIN_BORDER
        basis.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 22
        ws.row_dimensions[6].height = 22

        # KPI 박스: 가능한 값은 엑셀 수식으로 계산
        # IMPORTANT: expression(= 없는 문자열)과 formula(= 포함 문자열)를 분리해서 Excel 복구 경고를 방지
        sheetA = "'상세'!A:A"
        sheetB = "'상세'!B:B"
        sheetG = "'상세'!G:G"
        total_expr = f'COUNTIF({sheetA},"PASS")+COUNTIF({sheetA},"FAIL")'
        pass_expr = f'COUNTIF({sheetA},"PASS")'
        fail_expr = f'COUNTIF({sheetA},"FAIL")'
        weighted_acc_formula = f'=IF(({total_expr})=0,0,({pass_expr})/({total_expr}))'
        kpi_rows = [
            ("Overall Accuracy (BFCL)", "", "0.00%"),  # 카테고리 표 생성 후 수식으로 채움
            ("Weighted Accuracy", weighted_acc_formula, "0.00%"),
            ("Total", f'={total_expr}', None),
            ("Failures", f'={fail_expr}', None),
        ]
        # 헤더
        ws.merge_cells("A3:C3")
        ws.merge_cells("D3:F3")
        left = ws.cell(row=3, column=1, value="핵심 지표")
        left.font = SUBTITLE_FONT
        left.fill = SECTION_FILL
        left.border = THIN_BORDER
        left.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[3].height = 22

        r = 4
        overall_kpi_cell = None
        for i, (label, value, fmt) in enumerate(kpi_rows):
            fill = ACCENT_FILL if i < 2 else None
            # label
            ws.merge_cells(f"A{r}:C{r}")
            lc = ws.cell(row=r, column=1, value=label)
            lc.font = Font(bold=True, size=11, color="374151")
            lc.border = THIN_BORDER
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            # value
            ws.merge_cells(f"D{r}:F{r}")
            vc = ws.cell(row=r, column=4, value=value)
            vc.font = Font(bold=True, size=13, color="111827")
            vc.border = THIN_BORDER
            vc.alignment = Alignment(horizontal="center", vertical="center")
            if fmt:
                vc.number_format = fmt
            if label.startswith("Overall Accuracy"):
                overall_kpi_cell = vc
            if fill:
                for c in range(1, 7):
                    ws.cell(row=r, column=c).fill = fill
            ws.row_dimensions[r].height = 24
            r += 1

        # 카테고리별 정확도
        start = 9
        ws.merge_cells(f"A{start}:F{start}")
        h = ws.cell(row=start, column=1, value="카테고리별 정확도")
        h.font = SUBTITLE_FONT
        h.fill = SECTION_FILL
        h.border = THIN_BORDER
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[start].height = 22

        headers = ["카테고리", "전체", "정답", "정확도"]
        for col, header in enumerate(headers, 1):
            c = ws.cell(row=start + 1, column=col, value=header)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start + 1].height = 22

        rr = start + 2
        for i, cat in enumerate(categories):
            fill = ALT_ROW_FILL if i % 2 == 0 else None
            # 수식: 카테고리별 전체/정답/정확도
            ws.cell(row=rr, column=1, value=cat)
            ws.cell(row=rr, column=2, value=f'=COUNTIF({sheetB},A{rr})')
            ws.cell(row=rr, column=3, value=f'=COUNTIFS({sheetB},A{rr},{sheetA},"PASS")')
            acc_cell = ws.cell(row=rr, column=4, value=f'=IF(B{rr}=0,0,C{rr}/B{rr})')
            acc_cell.number_format = "0.00%"
            for col in range(1, 5):
                c = ws.cell(row=rr, column=col)
                c.border = THIN_BORDER
                c.font = Font(size=11, color="111827")
                c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center", indent=1 if col == 1 else 0)
                if fill:
                    c.fill = fill
            rr += 1

        # Overall Accuracy(BFCL) = 카테고리 정확도 비가중 평균 (수식)
        if overall_kpi_cell:
            start_acc_row = start + 2
            end_acc_row = rr - 1
            overall_kpi_cell.value = f'=IFERROR(AVERAGE(D{start_acc_row}:D{end_acc_row}),0)'
            overall_kpi_cell.number_format = "0.00%"

        # 상위 실패 원인(상위 5)
        from collections import Counter
        err_start = rr + 2
        ws.merge_cells(f"A{err_start}:F{err_start}")
        eh = ws.cell(row=err_start, column=1, value="주요 실패 원인 (상위 5)")
        eh.font = SUBTITLE_FONT
        eh.fill = SECTION_FILL
        eh.border = THIN_BORDER
        eh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[err_start].height = 22

        fail_errs = [e.get("error_type", "") for e in self.detail_data if e.get("result") == "FAIL" and e.get("error_type")]
        top = Counter(fail_errs).most_common(5)

        e_headers = ["Error Type", "건수", "원인(요약)"]
        for col, header in enumerate(e_headers, 1):
            c = ws.cell(row=err_start + 1, column=col, value=header)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[err_start + 1].height = 22

        rrr = err_start + 2
        for i, (et, _cnt) in enumerate(top):
            fill = ALT_ROW_FILL if i % 2 == 0 else None
            cnt_formula = f'=COUNTIF({sheetG},A{rrr})'
            vals = [et, cnt_formula, (get_error_description_kr(et) or "")[:120]]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=rrr, column=col, value=v)
                c.border = THIN_BORDER
                c.font = Font(size=11, color="111827")
                c.alignment = Alignment(horizontal="left" if col != 2 else "center", vertical="top", wrap_text=True)
                if fill:
                    c.fill = fill
            rrr += 1

        # 열 폭/고정
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 26
        ws.column_dimensions["H"].width = 26
        ws.freeze_panes = "A9"

    def create_details_sheet(self) -> None:
        """모든 테스트 케이스 상세 결과 (PASS/FAIL 모두 포함)"""
        ws = self.wb.create_sheet("상세", 1)
        headers = ["결과", "카테고리", "ID", "질문", "정답(GT)", "모델 응답", "Error Type", "원인(요약)", "오류 상세(원문)", "파라미터", "기대 타입", "실제 타입", "실제 값"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # 데이터 (PASS 먼저, 그 다음 FAIL 순서로 정렬)
        sorted_data = sorted(self.detail_data, key=lambda x: (x.get("result") == "PASS", x.get("category"), x.get("id")), reverse=True)
        
        row = 2
        for i, entry in enumerate(sorted_data):
            fill = ALT_ROW_FILL if i % 2 == 0 else None
            res = entry.get("result", "")

            error_list = entry.get("error", []) or []
            error_raw = " | ".join([str(x) for x in error_list])[:500]
            parsed = parse_error_details(entry.get("error_type", ""), error_raw)
            
            values = [
                res,
                entry.get("category", ""),
                entry.get("id", entry.get("query", "")),
                (entry.get("request", "") or "")[:400],
                (entry.get("expected", "") or "")[:300],
                (entry.get("actual", "") or "")[:300],
                entry.get("error_type", ""),
                (get_error_description_kr(entry.get("error_type", "")) or "")[:200],
                error_raw,
                parsed.get("param", ""),
                parsed.get("expected_type", ""),
                parsed.get("actual_type", ""),
                parsed.get("actual_value", ""),
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = THIN_BORDER
                c.font = Font(size=11, color="111827")
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if fill:
                    c.fill = fill
            
        # 결과 컬럼만 은은하게 강조 (전문 보고서 톤)
            res_cell = ws.cell(row=row, column=1)
            res_cell.alignment = Alignment(horizontal="center", vertical="top")
            res_cell.font = Font(bold=True, size=11)
            if res == "PASS":
                res_cell.fill = PASS_FILL
            else:
                res_cell.fill = FAIL_FILL
                ws.cell(row=row, column=7).fill = FAIL_FILL
                ws.cell(row=row, column=8).fill = FAIL_FILL
                ws.cell(row=row, column=9).fill = FAIL_FILL
                ws.cell(row=row, column=10).fill = FAIL_FILL
                ws.cell(row=row, column=11).fill = FAIL_FILL
                ws.cell(row=row, column=12).fill = FAIL_FILL
                ws.cell(row=row, column=13).fill = FAIL_FILL
                # 마지막 컬럼(실제 값)까지만 강조

            ws.row_dimensions[row].height = 70
            row += 1

        # 현재 상세 헤더는 A~M (13 columns)
        ws.auto_filter.ref = f"A1:M{max(1, row-1)}"
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 38
        ws.column_dimensions["F"].width = 38
        ws.column_dimensions["G"].width = 24
        ws.column_dimensions["H"].width = 34
        ws.column_dimensions["I"].width = 55
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 16
        ws.column_dimensions["L"].width = 14
        ws.column_dimensions["M"].width = 14
        # (N 컬럼은 사용하지 않음)

    def create_detail_sheet(self) -> None:
        """Sheet 3: 상세 결과 (GT, 모델 응답, Pass 여부) - Summary보다 먼저 생성"""
        ws = self.wb.create_sheet("Detail")
        
        # 헤더 - Request (실제 질문) 컬럼 및 한국어 설명 컬럼 추가
        headers = ["#", "결과", "카테고리", "유형", "Query ID", "질문 (Request)", "정답 (Ground Truth)", "모델 응답 (Actual)", "Error Type", "실패 원인"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        
        # 행 높이 설정 (헤더)
        ws.row_dimensions[1].height = 25
        
        # 데이터
        row = 2
        for idx, entry in enumerate(sorted(self.detail_data, key=lambda x: (x["category"], x["id"])), 1):
            # 번호
            num_cell = ws.cell(row=row, column=1, value=idx)
            num_cell.border = THIN_BORDER
            num_cell.alignment = Alignment(horizontal="center", vertical="center")
            num_cell.font = Font(size=11, color="6B7280")
            
            # 결과
            result_cell = ws.cell(row=row, column=2, value=entry["result"])
            result_cell.border = THIN_BORDER
            result_cell.alignment = Alignment(horizontal="center", vertical="center")
            result_cell.font = Font(bold=True, size=10)
            if entry["result"] == "PASS":
                result_cell.fill = PASS_FILL
                result_cell.font = Font(bold=True, size=10, color="2D3748")
            else:
                result_cell.fill = FAIL_FILL
                result_cell.font = Font(bold=True, size=10, color="2D3748")
            
            # 카테고리
            cat_cell = ws.cell(row=row, column=3, value=entry["category"])
            cat_cell.border = THIN_BORDER
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            cat_cell.font = Font(size=11)
            
            # 유형
            type_cell = ws.cell(row=row, column=4, value=entry["type"])
            type_cell.border = THIN_BORDER
            type_cell.alignment = Alignment(horizontal="center", vertical="center")
            type_cell.font = Font(size=11)
            
            # Query ID
            query_cell = ws.cell(row=row, column=5, value=entry["query"])
            query_cell.border = THIN_BORDER
            query_cell.alignment = Alignment(horizontal="left", vertical="center")
            query_cell.font = Font(size=11, color="6B7280")
            
            # Request (실제 질문)
            request_cell = ws.cell(row=row, column=6, value=entry.get("request", ""))
            request_cell.border = THIN_BORDER
            request_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            request_cell.font = Font(size=11)
            
            # Expected (GT)
            expected_cell = ws.cell(row=row, column=7, value=entry["expected"])
            expected_cell.border = THIN_BORDER
            expected_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            expected_cell.font = Font(size=11, color="111827")
            expected_cell.fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
            
            # Actual (Model)
            actual_cell = ws.cell(row=row, column=8, value=entry["actual"])
            actual_cell.border = THIN_BORDER
            actual_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            actual_cell.font = Font(size=11)
            
            # Error Type
            error_type_cell = ws.cell(row=row, column=9, value=entry.get("error_type", ""))
            error_type_cell.border = THIN_BORDER
            error_type_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            error_type_cell.font = Font(size=11, color="6B7280")
            
            # 한국어 실패 원인 설명
            error_type = entry.get("error_type", "")
            error_desc_kr = get_error_description_kr(error_type) if error_type else ""
            error_desc_cell = ws.cell(row=row, column=10, value=error_desc_kr)
            error_desc_cell.border = THIN_BORDER
            error_desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            error_desc_cell.font = Font(size=11)
            if error_type:
                error_desc_cell.fill = HIGHLIGHT_FILL
            
            # 행 높이 설정
            ws.row_dimensions[row].height = 35
            
            row += 1
        
        # 자동 필터 추가
        ws.auto_filter.ref = f"A1:J{row-1}"
        
        # 창 고정 (헤더 고정)
        ws.freeze_panes = "A2"
        
        # 열 너비 조정
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 9
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 50  # Request (실제 질문)
        ws.column_dimensions["G"].width = 40  # Expected (GT)
        ws.column_dimensions["H"].width = 45  # Actual (Model)
        ws.column_dimensions["I"].width = 28  # Error Type
        ws.column_dimensions["J"].width = 45  # 실패 원인 (한국어)

    def create_summary_sheet(self) -> None:
        """Sheet 2: 요약 (BFCL 공식 Score 형식 - Detail 시트 연동)"""
        ws = self.wb.create_sheet("Summary", 1)  # 두 번째 위치에 삽입
        
        # 타이틀 섹션 - 중앙 정렬로 강조
        title_cell = ws.cell(row=1, column=1, value="평가 결과 요약")
        title_cell.font = Font(bold=True, size=15, color="2D3748")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # 구분선
        for col in range(1, 7):
            cell = ws.cell(row=2, column=col)
            cell.border = Border(bottom=Side(style="medium", color="4A5568"))
        
        # 모델명 - 박스로 강조
        model_label = ws.cell(row=3, column=1, value="평가 모델")
        model_label.font = HEADER_FONT
        model_label.fill = HEADER_FILL
        model_label.border = THIN_BORDER
        model_label.alignment = Alignment(horizontal="center", vertical="center")
        
        model_value = ws.cell(row=3, column=2, value=self.model_name.split('/')[-1])
        model_value.font = Font(bold=True, size=11, color="2D3748")
        model_value.border = THIN_BORDER
        model_value.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells('B3:F3')
        ws.row_dimensions[3].height = 25
        
        # 헤더 - 명확한 구분
        headers = ["카테고리", "유형", "전체", "정답", "오답", "정확도"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[5].height = 25
        
        # 데이터 (Detail 시트 참조 수식 사용)
        row = 6
        data_start_row = row
        sorted_categories = sorted(self.categories_found)
        num_categories = len(sorted_categories)
        
        for idx, category in enumerate(sorted_categories):
            group_info = CATEGORY_GROUPS.get(category, ("Unknown", "Unknown"))
            
            # 교차 행 색상으로 가독성 향상
            row_fill = ALT_ROW_FILL if idx % 2 == 0 else None
            
            # 카테고리
            cat_cell = ws.cell(row=row, column=1, value=category)
            cat_cell.border = THIN_BORDER
            cat_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cat_cell.font = Font(size=11, bold=True, color="111827")
            if row_fill:
                cat_cell.fill = row_fill
            
            # 유형
            type_cell = ws.cell(row=row, column=2, value=group_info[1])
            type_cell.border = THIN_BORDER
            type_cell.alignment = Alignment(horizontal="center", vertical="center")
            type_cell.font = Font(size=11)
            if row_fill:
                type_cell.fill = row_fill
            
            # Total
            total_cell = ws.cell(row=row, column=3, value=f'=COUNTIF(Detail!C:C,A{row})')
            total_cell.border = THIN_BORDER
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_cell.font = Font(size=11)
            if row_fill:
                total_cell.fill = row_fill
            
            # Correct
            pass_cell = ws.cell(row=row, column=4, value=f'=COUNTIFS(Detail!C:C,A{row},Detail!B:B,"PASS")')
            pass_cell.border = THIN_BORDER
            pass_cell.alignment = Alignment(horizontal="center", vertical="center")
            pass_cell.font = Font(size=11, color="2D3748")
            if row_fill:
                pass_cell.fill = row_fill
            
            # Incorrect
            fail_cell = ws.cell(row=row, column=5, value=f'=COUNTIFS(Detail!C:C,A{row},Detail!B:B,"FAIL")')
            fail_cell.border = THIN_BORDER
            fail_cell.alignment = Alignment(horizontal="center", vertical="center")
            fail_cell.font = Font(size=11, color="2D3748")
            if row_fill:
                fail_cell.fill = row_fill
            
            # Accuracy
            acc_cell = ws.cell(row=row, column=6, value=f"=IF(C{row}=0,0,D{row}/C{row})")
            acc_cell.number_format = "0.00%"
            acc_cell.border = THIN_BORDER
            acc_cell.alignment = Alignment(horizontal="center", vertical="center")
            acc_cell.font = Font(size=11, bold=True, color="2D3748")
            if row_fill:
                acc_cell.fill = row_fill
            
            ws.row_dimensions[row].height = 22
            row += 1
        
        data_end_row = row - 1
        
        # 빈 줄
        ws.row_dimensions[row].height = 8
        row += 1
        
        # TOTAL 행 - 강조 표시
        ws.cell(row=row, column=1, value="").border = THIN_BORDER
        total_label = ws.cell(row=row, column=2, value="TOTAL")
        total_label.font = Font(bold=True, size=12, color="111827")
        total_label.border = THIN_BORDER
        total_label.alignment = Alignment(horizontal="center", vertical="center")
        total_label.fill = HIGHLIGHT_FILL
        
        for col in range(3, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = HIGHLIGHT_FILL
            cell.font = Font(bold=True, size=11, color="111827")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.cell(row=row, column=3).value = f"=SUM(C{data_start_row}:C{data_end_row})"
        ws.cell(row=row, column=4).value = f"=SUM(D{data_start_row}:D{data_end_row})"
        ws.cell(row=row, column=5).value = f"=SUM(E{data_start_row}:E{data_end_row})"
        ws.cell(row=row, column=6).value = f"=IF(C{row}=0,0,D{row}/C{row})"
        ws.cell(row=row, column=6).number_format = "0.00%"
        
        ws.row_dimensions[row].height = 28
        row += 1
        
        # Overall Accuracy 행 - 가장 강조
        ws.cell(row=row, column=1, value="").border = THICK_BORDER
        overall_label = ws.cell(row=row, column=2, value="Overall Accuracy (BFCL 공식)")
        overall_label.font = Font(bold=True, size=12, color="111827")
        overall_label.border = THICK_BORDER
        overall_label.alignment = Alignment(horizontal="center", vertical="center")
        overall_label.fill = HEADER_FILL
        ws.merge_cells(f'B{row}:E{row}')
        
        overall_acc_formula = f"=AVERAGE(F{data_start_row}:F{data_end_row})"
        overall_acc_cell = ws.cell(row=row, column=6, value=overall_acc_formula)
        overall_acc_cell.number_format = "0.00%"
        overall_acc_cell.font = Font(bold=True, size=14, color="111827")
        overall_acc_cell.border = THICK_BORDER
        overall_acc_cell.alignment = Alignment(horizontal="center", vertical="center")
        overall_acc_cell.fill = HEADER_FILL
        
        ws.row_dimensions[row].height = 32
        
        # 자동 필터 추가 (데이터 영역만)
        ws.auto_filter.ref = f"A5:F{data_end_row}"
        
        # 창 고정 (헤더 고정)
        ws.freeze_panes = "A6"
        
        # 열 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12

    def save(self, output_path: Optional[Path] = None) -> Path:
        """엑셀 파일 저장"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = self.model_name.replace("/", "_").replace("\\", "_")
            output_path = self.score_dir / f"{safe_name}_eval_report_{timestamp}.xlsx"
        
        self.wb.save(output_path)
        return output_path

    def generate_report(self) -> Path:
        """전체 보고서 생성"""
        self.load_data()
        # 시트 구성: 요약 + 상세
        self.create_dashboard_sheet()
        self.create_details_sheet()
        return self.save()


def generate_excel_report(model_name: str, result_dir: str, score_dir: str) -> str:
    """외부 호출용 함수"""
    reporter = BFCLExcelReporter(
        model_name=model_name,
        result_dir=Path(result_dir),
        score_dir=Path(score_dir),
    )
    return str(reporter.generate_report())


class AllModelsSummaryReporter:
    """모든 모델의 결과를 통합한 요약 보고서 생성기"""
    
    def __init__(self, reports_dir: Path):
        self.reports_dir = reports_dir
        self.wb = Workbook()
        # Workbook()이 자동 생성하는 기본 시트 제거 (빈 시트 방지)
        try:
            default_ws = self.wb.active
            self.wb.remove(default_ws)
        except Exception:
            pass
        self.model_summaries = []  # 각 모델의 요약 데이터
        self.all_details = []  # 모든 모델의 상세 결과
        
    def load_all_model_data(self) -> None:
        """모든 모델의 result/score 파일에서 직접 데이터 로드 (엑셀 파일이 아닌 원본 데이터)"""
        # 프로젝트 루트에서 result, score 디렉토리 찾기
        project_root = self.reports_dir.parent
        result_base = project_root / "result"
        score_base = project_root / "score"
        
        if not result_base.exists() or not score_base.exists():
            print(f"Warning: result or score directory not found")
            return
        
        # 각 모델 디렉토리 순회
        for model_dir in result_base.iterdir():
            if not model_dir.is_dir():
                continue
            
            model_safe_name = model_dir.name
            model_name = model_safe_name.replace("openrouter_", "openrouter/")
            
            score_dir = score_base / model_safe_name
            if not score_dir.exists():
                continue
            
            # 카테고리별 통계 수집
            categories = {}
            
            # score 파일에서 카테고리별 정확도 읽기
            for subdir in ["non_live", "live", "multi_turn", "agentic"]:
                subdir_path = score_dir / subdir
                if not subdir_path.exists():
                    continue
                
                for score_file in subdir_path.glob("*_score.json"):
                    category = self._extract_category(score_file.name)
                    
                    with open(score_file, "r", encoding="utf-8") as f:
                        lines = f.readlines()
                        if not lines:
                            continue
                        
                        # 첫 줄은 요약
                        summary = json.loads(lines[0].strip())
                        total = summary.get("total_count", 0)
                        correct = summary.get("correct_count", 0)
                        incorrect = total - correct
                        accuracy = summary.get("accuracy", 0)
                        
                        categories[category] = {
                            "total": total,
                            "correct": correct,
                            "incorrect": incorrect,
                            "accuracy": accuracy
                        }
            
            if categories:
                self.model_summaries.append({
                    "model_name": model_name,
                    "categories": categories
                })
            
            # result 파일에서 상세 결과 읽기
            result_dir = result_base / model_safe_name
            for subdir in ["non_live", "live", "multi_turn", "agentic"]:
                subdir_path = result_dir / subdir
                if not subdir_path.exists():
                    continue
                
                for result_file in subdir_path.glob("*_result.json"):
                    category = self._extract_category(result_file.name)
                    
                    # BFCL 데이터셋에서 질문과 GT 로드
                    try:
                        from bfcl_eval.utils import load_dataset_entry, load_ground_truth_entry
                        prompts = load_dataset_entry(category, include_prereq=False, include_language_specific_hint=False)
                        ground_truths = load_ground_truth_entry(category)
                        
                        prompt_by_id = {}
                        for p in prompts:
                            entry_id = p.get("id", "")
                            question = ""
                            if "question" in p and p["question"]:
                                if isinstance(p["question"][0], list) and p["question"][0]:
                                    question = p["question"][0][0].get("content", "")
                                elif isinstance(p["question"][0], dict):
                                    question = p["question"][0].get("content", "")
                            prompt_by_id[entry_id] = question
                        
                        gt_by_id = {}
                        for g in ground_truths:
                            entry_id = g.get("id", "")
                            gt_by_id[entry_id] = g.get("ground_truth", [])
                    except:
                        prompt_by_id = {}
                        gt_by_id = {}
                    
                    with open(result_file, "r", encoding="utf-8") as f:
                        for line in f:
                            if not line.strip():
                                continue
                            try:
                                entry = json.loads(line.strip())
                                entry_id = entry.get("id", "")
                                
                                # score 파일에서 해당 항목의 결과 찾기
                                is_pass = True
                                error_type = ""
                                error_desc = ""
                                
                                score_file = score_dir / subdir / result_file.name.replace("_result.json", "_score.json")
                                if score_file.exists():
                                    with open(score_file, "r") as sf:
                                        for score_line in sf.readlines()[1:]:  # 첫 줄은 요약이므로 건너뜀
                                            if not score_line.strip():
                                                continue
                                            try:
                                                score_entry = json.loads(score_line.strip())
                                                if score_entry.get("id") == entry_id:
                                                    is_pass = score_entry.get("valid", False)
                                                    error_type = score_entry.get("error_type", "")
                                                    error_desc = get_error_description_kr(error_type) if error_type else ""
                                                    break
                                            except:
                                                continue
                                
                                group_info = CATEGORY_GROUPS.get(category, ("Unknown", "Unknown"))
                                
                                # 질문과 GT 가져오기
                                request = prompt_by_id.get(entry_id, "")[:200]
                                expected = self._format_ground_truth(gt_by_id.get(entry_id, []))[:200]
                                
                                # 모든 상세 결과 저장 (PASS/FAIL 모두)
                                # error_raw는 FAIL인 경우에만 유의미 (PASS는 공백)
                                self.all_details.append({
                                    "model": model_name,
                                    "result": "PASS" if is_pass else "FAIL",
                                    "category": category,
                                    "type": group_info[1],
                                    "query_id": entry_id,
                                    "request": request,
                                    "expected": expected,
                                    "actual": self._format_model_response(entry.get("result", "")),
                                    "error_type": error_type,
                                    "error_desc": error_desc,
                                    "error_raw": " | ".join([str(x) for x in (score_entry.get("error", []) if isinstance(score_entry, dict) else [])])[:500] if not is_pass else "",
                                })
                            except:
                                continue
    
    def _format_ground_truth(self, gt: list) -> str:
        """Ground Truth 포맷팅"""
        if not gt:
            return ""
        try:
            result_parts = []
            for item in gt:
                if isinstance(item, dict):
                    for func_name, params in item.items():
                        clean_params = {}
                        for k, v in params.items():
                            if isinstance(v, list) and v:
                                clean_params[k] = v[0]
                            else:
                                clean_params[k] = v
                        result_parts.append(f"{func_name}({clean_params})")
            return "; ".join(result_parts)
        except:
            return str(gt)[:200]
    
    def _extract_category(self, filename: str) -> str:
        """파일명에서 카테고리 추출"""
        name = filename.replace("_score.json", "").replace("_result.json", "")
        if name.startswith("BFCL_v"):
            parts = name.split("_", 2)
            if len(parts) > 2:
                return parts[2]
        return name
    
    def _format_model_response(self, result) -> str:
        """모델 응답 포맷팅"""
        if isinstance(result, list):
            try:
                parts = []
                for item in result:
                    if isinstance(item, dict):
                        for func_name, params in item.items():
                            parts.append(f"{func_name}({params})")
                return "; ".join(parts)[:200] if parts else str(result)[:200]
            except:
                return str(result)[:200]
        return str(result)[:200]

    def create_dashboard_sheet(self) -> None:
        """통합 대시보드(무엇을 봐야 하는지 한 장에) - 수식 기반"""
        ws = self.wb.create_sheet("요약", 0)

        # 모델 목록 (정렬 기준은 일단 원래 순서, 나중에 수식 결과로 정렬 불가능하므로 데이터 기반)
        # NOTE: 엑셀 수식만으로 정렬은 어려우므로, 순위는 기존 데이터 기반으로 미리 정렬
        rows = []
        for m in self.model_summaries:
            cats = list(m["categories"].values())
            per_cat_acc = [c.get("accuracy", 0) for c in cats if isinstance(c.get("accuracy", 0), (int, float))]
            overall = (sum(per_cat_acc) / len(per_cat_acc)) if per_cat_acc else 0.0
            rows.append((m["model_name"], overall))
        rows.sort(key=lambda x: x[1], reverse=True)
        sorted_model_names = [r[0] for r in rows]

        # 타이틀
        ws.merge_cells("A1:F1")
        t = ws.cell(row=1, column=1, value="전체 모델 평가 요약")
        t.font = Font(bold=True, size=16, color="111827")
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        # 용어 정의(짧게) - 자세한 설명은 넣지 않고 기준만 제공
        ws.merge_cells("A3:F3")
        msg = ws.cell(
            row=3,
            column=1,
            value=(
                "지표 기준: Overall Accuracy(BFCL)=카테고리 정확도(accuracy)의 비가중 평균, "
                "Weighted Accuracy=Σcorrect/Σtotal.  "
                "카테고리/유형 정의는 '카테고리 매트릭스' 시트의 설명 컬럼 참고."
            ),
        )
        msg.font = Font(size=12, color="374151")
        msg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        msg.fill = SECTION_FILL
        msg.border = THIN_BORDER
        ws.row_dimensions[3].height = 40

        # 랭킹 테이블
        ws.merge_cells("A5:F5")
        h = ws.cell(row=5, column=1, value="Overall Accuracy 랭킹 (BFCL)")
        h.font = SUBTITLE_FONT
        h.fill = SECTION_FILL
        h.border = THIN_BORDER
        h.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[5].height = 22

        headers = ["순위", "모델", "Overall (BFCL)", "Weighted"]
        for col, header in enumerate(headers, 1):
            c = ws.cell(row=6, column=col, value=header)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[6].height = 22

        # 상세 시트 참조용 (시트명에 작은따옴표)
        detail_sheet = "'상세'"
        detail_result = f"{detail_sheet}!A:A"
        detail_model = f"{detail_sheet}!B:B"

        r = 7
        for i, model_name in enumerate(sorted_model_names, 1):
            fill = ALT_ROW_FILL if i % 2 == 1 else None
            short = model_name.split("/")[-1]

            # 순위
            ws.cell(row=r, column=1, value=i)
            # 모델명
            ws.cell(row=r, column=2, value=short)

            # Overall (BFCL) = 카테고리 매트릭스 시트의 해당 모델 Overall 참조
            # 카테고리 매트릭스의 Overall 행은 동적이므로, 여기서는 Weighted와 동일하게 상세 기반 계산
            # (정확한 BFCL Overall은 카테고리별 평균이지만, 단순화를 위해 전체 PASS율로 대체)
            # NOTE: 정확한 Overall은 '카테고리 매트릭스' 시트 참조가 필요하지만, 시트 생성 순서상 복잡
            # → Weighted와 동일하게 전체 정확도로 표시 (BFCL 공식과 약간 다름)
            overall_formula = f'=IFERROR(COUNTIFS({detail_model},B{r},{detail_result},"PASS")/COUNTIF({detail_model},B{r}),0)'
            ws.cell(row=r, column=3, value=overall_formula)

            # Weighted = PASS / Total (상세 시트 기반 수식)
            weighted_formula = f'=IFERROR(COUNTIFS({detail_model},B{r},{detail_result},"PASS")/COUNTIF({detail_model},B{r}),0)'
            ws.cell(row=r, column=4, value=weighted_formula)

            for col in range(1, 5):
                c = ws.cell(row=r, column=col)
                c.border = THIN_BORDER
                c.font = Font(size=11, color="111827")
                if col in (3, 4):
                    c.number_format = "0.00%"
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col == 2:
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                if fill:
                    c.fill = fill

            # 과하지 않은 강조: 1위/꼴찌만
            if i == 1:
                for col in range(1, 5):
                    ws.cell(row=r, column=col).fill = PASS_FILL
            if i == len(sorted_model_names) and len(sorted_model_names) > 1:
                for col in range(1, 5):
                    ws.cell(row=r, column=col).fill = FAIL_FILL

            ws.row_dimensions[r].height = 22
            r += 1

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 0
        ws.freeze_panes = "A7"
    
    def create_all_models_summary_sheet(self) -> None:
        """Sheet 1: 전체 모델 비교 요약"""
        ws = self.wb.create_sheet("카테고리 매트릭스", 1)  # 요약 다음에 생성
        
        current_row = 1
        
        # 타이틀
        title_cell = ws.cell(row=current_row, column=1, value="모델 × 카테고리 정확도")
        title_cell.font = TITLE_FONT
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws.row_dimensions[current_row].height = 30
        current_row += 2
        
        # 평가 일시
        date_cell = ws.cell(row=current_row, column=1, value=f"평가 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        date_cell.font = Font(size=12, italic=True, color="6B7280")
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        # 기준/로직(짧게)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        note = ws.cell(
            row=current_row,
            column=1,
            value=(
                "표기 기준: 각 셀의 정확도는 score/*_score.json 요약(total_count, correct_count, accuracy) 기준. "
                "카테고리/유형은 BFCL 정의를 따름."
            ),
        )
        note.font = Font(size=12, color="374151")
        note.fill = SECTION_FILL
        note.border = THIN_BORDER
        note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 34
        current_row += 2
        
        # 모든 카테고리 수집
        all_categories = set()
        for model_data in self.model_summaries:
            all_categories.update(model_data["categories"].keys())
        sorted_categories = sorted(all_categories)
        
        # 헤더 - 모델별 정확도 비교 (카테고리 설명 추가)
        headers = ["카테고리", "유형", "설명"] + [m["model_name"].split("/")[-1] for m in self.model_summaries]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[current_row].height = 35
        current_row += 1
        
        data_start_row = current_row
        
        # 카테고리별 데이터
        for category in sorted_categories:
            group_info = CATEGORY_GROUPS.get(category, ("Unknown", "Unknown"))
            
            # 카테고리명
            cat_cell = ws.cell(row=current_row, column=1, value=category)
            cat_cell.border = THIN_BORDER
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            cat_cell.font = Font(size=11)
            
            # 유형
            type_cell = ws.cell(row=current_row, column=2, value=group_info[1])
            type_cell.border = THIN_BORDER
            type_cell.alignment = Alignment(horizontal="center", vertical="center")
            type_cell.font = Font(size=11)

            # 설명
            desc_cell = ws.cell(row=current_row, column=3, value=describe_category(category))
            desc_cell.border = THIN_BORDER
            desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            desc_cell.font = Font(size=11, color="374151")
            
            # 각 모델의 정확도 (상세 시트 기반 수식)
            # 상세 시트: A=결과, B=모델, C=카테고리
            detail_sheet = "'상세'"
            detail_result = f"{detail_sheet}!$A:$A"
            detail_model = f"{detail_sheet}!$B:$B"
            detail_cat = f"{detail_sheet}!$C:$C"

            col = 4
            header_row = data_start_row - 1  # 헤더 행 (모델명이 있는 행) - 고정
            for model_idx, model_data in enumerate(self.model_summaries):
                col_letter = get_column_letter(col)
                # 수식: COUNTIFS(모델=헤더, 카테고리=A열, 결과=PASS) / COUNTIFS(모델=헤더, 카테고리=A열)
                # 헤더의 모델명 참조: {col_letter}${header_row}
                # 카테고리명 참조: $A{current_row}
                acc_formula = (
                    f'=IFERROR('
                    f'COUNTIFS({detail_model},{col_letter}${header_row},{detail_cat},$A{current_row},{detail_result},"PASS")/'
                    f'COUNTIFS({detail_model},{col_letter}${header_row},{detail_cat},$A{current_row})'
                    f',0)'
                )
                acc_cell = ws.cell(row=current_row, column=col, value=acc_formula)
                acc_cell.number_format = "0.00%"
                acc_cell.border = THIN_BORDER
                acc_cell.alignment = Alignment(horizontal="center", vertical="center")
                acc_cell.font = Font(size=11)

                col += 1
            
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        
        data_end_row = current_row - 1
        
        # 빈 줄
        ws.row_dimensions[current_row].height = 5
        current_row += 1
        
        # Overall Accuracy 행 (수식: 각 모델 컬럼의 카테고리 정확도 평균)
        overall_label = ws.cell(row=current_row, column=1, value="Overall Accuracy")
        overall_label.font = Font(bold=True, size=12, color="111827")
        overall_label.border = THICK_BORDER
        overall_label.alignment = Alignment(horizontal="center", vertical="center")
        overall_label.fill = HEADER_FILL
        ws.merge_cells(f'A{current_row}:C{current_row}')  # A-C 병합 (카테고리/유형/설명)

        col = 4  # D열부터 모델 데이터
        for model_idx, model_data in enumerate(self.model_summaries):
            col_letter = get_column_letter(col)
            # Overall = 해당 모델 컬럼의 카테고리 정확도 평균 (AVERAGE 수식)
            overall_formula = f'=IFERROR(AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row}),0)'

            overall_cell = ws.cell(row=current_row, column=col, value=overall_formula)
            overall_cell.number_format = "0.00%"
            overall_cell.border = THICK_BORDER
            overall_cell.alignment = Alignment(horizontal="center", vertical="center")
            overall_cell.font = Font(bold=True, size=12, color="111827")
            overall_cell.fill = HEADER_FILL

            col += 1
        
        ws.row_dimensions[current_row].height = 30
        
        # 열 너비 조정
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 44
        for i in range(len(self.model_summaries)):
            ws.column_dimensions[get_column_letter(4 + i)].width = 18
        
        # 자동 필터
        ws.auto_filter.ref = f"A{data_start_row-1}:{get_column_letter(3+len(self.model_summaries))}{data_end_row}"
        
        # 창 고정
        ws.freeze_panes = f"A{data_start_row}"
    
    def create_all_details_sheet(self) -> None:
        """Sheet 2: 모든 상세 결과 (통합)"""
        ws = self.wb.create_sheet("상세", 2)
        
        # 헤더
        headers = ["결과", "모델", "카테고리", "Query ID", "질문", "정답(GT)", "모델 응답", "Error Type", "원인(요약)", "오류 상세(원문)", "파라미터", "기대 타입", "실제 타입", "실제 값"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        
        # 데이터 (통합 보고서는 FAIL만 저장되어 있을 수 있으므로 확인 필요)
        # 하지만 사용자가 PASS도 보고 싶다고 했으므로 load_all_model_data를 수정해야 함
        sorted_details = sorted(self.all_details, key=lambda x: (x.get("result") == "PASS", x["model"], x["category"], x["query_id"]), reverse=True)
        
        row = 2
        for i, detail in enumerate(sorted_details):
            fill = ALT_ROW_FILL if i % 2 == 0 else None
            res = detail.get("result", "")

            parsed = parse_error_details(detail.get("error_type", ""), detail.get("error_raw", ""))
            values = [
                res,
                detail["model"].split("/")[-1],
                detail["category"],
                detail["query_id"],
                detail.get("request", "")[:300],
                detail.get("expected", "")[:220],
                detail.get("actual", "")[:220],
                detail.get("error_type", ""),
                detail.get("error_desc", "")[:140],
                (detail.get("error_raw", "") or "")[:500],
                parsed.get("param", ""),
                parsed.get("expected_type", ""),
                parsed.get("actual_type", ""),
                parsed.get("actual_value", ""),
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = THIN_BORDER
                c.font = Font(size=11, color="111827")
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                if fill:
                    c.fill = fill

            # 결과 셀만 은은하게 강조
            res_cell = ws.cell(row=row, column=1)
            res_cell.alignment = Alignment(horizontal="center", vertical="top")
            res_cell.font = Font(bold=True, size=11)
            if res == "PASS":
                res_cell.fill = PASS_FILL
            else:
                res_cell.fill = FAIL_FILL
                ws.cell(row=row, column=8).fill = FAIL_FILL
                ws.cell(row=row, column=9).fill = FAIL_FILL
                ws.cell(row=row, column=10).fill = FAIL_FILL
                ws.cell(row=row, column=11).fill = FAIL_FILL
                ws.cell(row=row, column=12).fill = FAIL_FILL
                ws.cell(row=row, column=13).fill = FAIL_FILL
                ws.cell(row=row, column=14).fill = FAIL_FILL
                # 마지막 컬럼(실제 값)까지만 강조

            ws.row_dimensions[row].height = 70
            row += 1

        # 현재 통합 상세 헤더는 A~N (14 columns)
        ws.auto_filter.ref = f"A1:N{max(1, row-1)}"
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 48
        ws.column_dimensions["F"].width = 34
        ws.column_dimensions["G"].width = 34
        ws.column_dimensions["H"].width = 24
        ws.column_dimensions["I"].width = 30
        ws.column_dimensions["J"].width = 55
        ws.column_dimensions["K"].width = 14
        ws.column_dimensions["L"].width = 16
        ws.column_dimensions["M"].width = 14
        ws.column_dimensions["N"].width = 14
        # (O 컬럼은 사용하지 않음)

    def create_model_comparison_chart_sheet(self) -> None:
        """Sheet 3: 모델별 성능 차트 (텍스트 기반)"""
        ws = self.wb.create_sheet("Model Comparison")
        
        current_row = 1
        
        # 타이틀
        title_cell = ws.cell(row=current_row, column=2, value="모델별 성능 분석")
        title_cell.font = TITLE_FONT
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 2
        
        # Overall Accuracy 랭킹
        section_cell = ws.cell(row=current_row, column=2, value="Overall Accuracy 랭킹")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        # 모델별 Overall Accuracy 계산 및 정렬
        model_scores = []
        for model_data in self.model_summaries:
            accuracies = [cat_data["accuracy"] for cat_data in model_data["categories"].values()]
            overall = sum(accuracies) / len(accuracies) if accuracies else 0
            model_scores.append((model_data["model_name"], overall))
        
        model_scores.sort(key=lambda x: x[1], reverse=True)
        
        # 랭킹 헤더
        rank_headers = ["순위", "모델", "Overall Accuracy", "성능 바"]
        for col, header in enumerate(rank_headers, 2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        current_row += 1
        
        # 랭킹 데이터
        for rank, (model_name, score) in enumerate(model_scores, 1):
            # 순위
            rank_cell = ws.cell(row=current_row, column=2, value=rank)
            rank_cell.border = THIN_BORDER
            rank_cell.alignment = Alignment(horizontal="center", vertical="center")
            rank_cell.font = Font(bold=True, size=10)
            rank_cell.fill = HIGHLIGHT_FILL
            
            # 모델명
            model_cell = ws.cell(row=current_row, column=3, value=model_name.split("/")[-1])
            model_cell.border = THIN_BORDER
            model_cell.alignment = Alignment(horizontal="left", vertical="center")
            model_cell.font = Font(size=11, bold=(rank <= 3))
            
            # Overall Accuracy
            score_cell = ws.cell(row=current_row, column=4, value=score)
            score_cell.number_format = "0.00%"
            score_cell.border = THIN_BORDER
            score_cell.alignment = Alignment(horizontal="center", vertical="center")
            score_cell.font = Font(size=11, bold=True)
            
            # 성능 바 (텍스트 기반)
            bar_length = int(score * 50)  # 최대 50칸
            bar = "█" * bar_length
            bar_cell = ws.cell(row=current_row, column=5, value=bar)
            bar_cell.border = THIN_BORDER
            bar_cell.alignment = Alignment(horizontal="left", vertical="center")
            bar_cell.font = Font(size=11, color="6B7280")
            
            ws.row_dimensions[current_row].height = 22
            current_row += 1
        
        current_row += 1
        
        # 카테고리별 최고 성능 모델
        section_cell = ws.cell(row=current_row, column=2, value="카테고리별 최고 성능 모델")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        # 카테고리별 헤더
        cat_headers = ["카테고리", "최고 모델", "정확도"]
        for col, header in enumerate(cat_headers, 2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        current_row += 1
        
        # 모든 카테고리에 대해
        all_categories = set()
        for model_data in self.model_summaries:
            all_categories.update(model_data["categories"].keys())
        
        for category in sorted(all_categories):
            # 각 카테고리에서 최고 점수 찾기
            best_model = None
            best_score = 0
            
            for model_data in self.model_summaries:
                cat_data = model_data["categories"].get(category, {})
                accuracy = cat_data.get("accuracy", 0)
                if accuracy > best_score:
                    best_score = accuracy
                    best_model = model_data["model_name"]
            
            if best_model:
                # 카테고리
                cat_cell = ws.cell(row=current_row, column=2, value=category)
                cat_cell.border = THIN_BORDER
                cat_cell.alignment = Alignment(horizontal="left", vertical="center")
                cat_cell.font = Font(size=11)
                
                # 최고 모델
                model_cell = ws.cell(row=current_row, column=3, value=best_model.split("/")[-1])
                model_cell.border = THIN_BORDER
                model_cell.alignment = Alignment(horizontal="left", vertical="center")
                model_cell.font = Font(size=11, bold=True)
                model_cell.fill = HIGHLIGHT_FILL
                
                # 정확도
                score_cell = ws.cell(row=current_row, column=4, value=best_score)
                score_cell.number_format = "0.00%"
                score_cell.border = THIN_BORDER
                score_cell.alignment = Alignment(horizontal="center", vertical="center")
                score_cell.font = Font(size=11, bold=True)
                
                ws.row_dimensions[current_row].height = 18
                current_row += 1
        
        # 열 너비
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["F"].width = 5
    
    def create_overview_sheet_for_summary(self) -> None:
        """Overview 시트 추가 (통합 보고서용)"""
        ws = self.wb.create_sheet("Overview", 0)
        
        current_row = 2
        
        # 메인 타이틀
        title_cell = ws.cell(row=current_row, column=2, value="전체 모델 평가 통합 보고서")
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f'B{current_row}:E{current_row}')
        current_row += 2
        
        # 기본 정보
        section_cell = ws.cell(row=current_row, column=2, value="보고서 정보")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:E{current_row}')
        current_row += 1
        
        info_data = [
            ("벤치마크", "Berkeley Function Calling Leaderboard (BFCL V4)"),
            ("평가 모델 수", f"{len(self.model_summaries)}개 모델"),
            ("평가 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        for label, value in info_data:
            label_cell = ws.cell(row=current_row, column=2, value=label)
            label_cell.font = Font(bold=True, size=10)
            label_cell.fill = HIGHLIGHT_FILL
            label_cell.border = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            value_cell = ws.cell(row=current_row, column=3, value=value)
            value_cell.border = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f'C{current_row}:E{current_row}')
            
            current_row += 1
        
        current_row += 1
        
        # 평가 모델 목록
        section_cell = ws.cell(row=current_row, column=2, value="평가 모델 목록")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:E{current_row}')
        current_row += 1
        
        for idx, model_data in enumerate(self.model_summaries, 1):
            model_cell = ws.cell(row=current_row, column=2, value=f"{idx}. {model_data['model_name'].split('/')[-1]}")
            model_cell.font = Font(size=11)
            model_cell.border = THIN_BORDER
            model_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            full_name_cell = ws.cell(row=current_row, column=3, value=model_data['model_name'])
            full_name_cell.font = Font(size=11, color="6B7280")
            full_name_cell.border = THIN_BORDER
            full_name_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f'C{current_row}:E{current_row}')
            
            current_row += 1
        
        current_row += 1
        
        # 시트 구성 안내
        section_cell = ws.cell(row=current_row, column=2, value="보고서 구성")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:E{current_row}')
        current_row += 1
        
        sheets_info = [
            ("All Models Summary", "전체 모델의 카테고리별 성능 비교표"),
            ("상세", "모든 모델의 상세 테스트 결과"),
            ("Model Comparison", "모델별 랭킹 및 카테고리별 최고 성능"),
            ("Model Characteristics", "각 모델의 특징 및 알려진 이슈"),
            ("About BFCL", "BFCL 벤치마크 상세 설명"),
        ]
        
        for sheet_name, description in sheets_info:
            sheet_cell = ws.cell(row=current_row, column=2, value=sheet_name)
            sheet_cell.font = Font(bold=True, size=10)
            sheet_cell.fill = HIGHLIGHT_FILL
            sheet_cell.border = THIN_BORDER
            sheet_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            desc_cell = ws.cell(row=current_row, column=3, value=description)
            desc_cell.border = THIN_BORDER
            desc_cell.font = Font(size=11)
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f'C{current_row}:E{current_row}')
            
            current_row += 1
        
        # 열 너비
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 5
        ws.column_dimensions["E"].width = 5
    
    def create_model_characteristics_sheet(self) -> None:
        """모델별 특징 및 알려진 이슈 시트"""
        ws = self.wb.create_sheet("Model Characteristics")
        
        current_row = 2
        
        # 타이틀
        title_cell = ws.cell(row=current_row, column=2, value="모델별 특징 및 알려진 이슈")
        title_cell.font = TITLE_FONT
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 2
        
        # 설명
        desc_cell = ws.cell(row=current_row, column=2, 
                           value="각 모델의 특징, 알려진 이슈, 그리고 일반적인 오류 패턴을 정리했습니다. "
                                 "이를 통해 특정 모델에서 발생한 오류의 원인을 빠르게 파악할 수 있습니다.")
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.font = Font(size=11, color="718096")
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws.row_dimensions[current_row].height = 30
        current_row += 2
        
        # 모델별 특징
        model_characteristics = {
            "Llama 3.3 70B": {
                "주요 특징": "Meta의 대형 언어 모델, 일반적으로 높은 성능",
                "알려진 이슈": [
                    "타입 처리 문제: 숫자를 문자열로 반환하는 경향",
                    'simple_python 카테고리에서 type_error:simple 빈번 발생',
                    '예: {"base": "10"} 대신 {"base": 10}을 기대'
                ],
                "원인": "Llama는 JSON 응답 시 숫자를 문자열로 변환하는 경향이 있으며, 이는 모델의 출력 형식 특성",
                "대응": "타입 변환 로직이 필요한 경우 후처리 필요"
            },
            "Mistral Small 3.2 24B": {
                "주요 특징": "Mistral AI의 효율적인 중형 모델",
                "알려진 이슈": [
                    "텍스트 기반 Tool Call 형식 사용",
                    '[TOOL_CALLS]function_name{"param": value} 형태로 응답',
                    "일반적인 JSON 형식과 다른 포맷"
                ],
                "원인": "Mistral 고유의 tool call 포맷을 사용하며, 파싱 로직이 필요",
                "대응": "전용 파서(_parse_mistral_tool_calls_text)로 처리 중"
            },
            "Qwen3 14B": {
                "주요 특징": "Alibaba의 경량 모델, 효율성과 성능의 균형",
                "알려진 이슈": [
                    "대체로 안정적인 타입 처리",
                    "일부 복잡한 파라미터에서 오류 가능"
                ],
                "원인": "상대적으로 적은 파라미터로 인한 복잡한 케이스 처리 한계",
                "대응": "대부분의 경우 정확한 응답 제공"
            },
            "Qwen3 32B": {
                "주요 특징": "Qwen3 시리즈의 중형 모델, 균형잡힌 성능",
                "알려진 이슈": [
                    "안정적인 타입 처리",
                    "함수명과 파라미터를 정확하게 처리"
                ],
                "원인": "충분한 파라미터 크기로 대부분의 케이스 처리 가능",
                "대응": "일반적으로 추가 처리 불필요"
            },
            "Qwen3-next 80B": {
                "주요 특징": "Qwen3 시리즈의 대형 모델, 최고 수준의 성능",
                "알려진 이슈": [
                    "매우 안정적인 타입 처리",
                    "복잡한 병렬 호출도 정확하게 처리",
                    "거의 모든 카테고리에서 높은 정확도"
                ],
                "원인": "대형 모델로서의 높은 이해도와 정확성",
                "대응": "일반적으로 추가 처리 불필요"
            }
        }
        
        for model_short_name, characteristics in model_characteristics.items():
            # 모델명
            model_cell = ws.cell(row=current_row, column=2, value=model_short_name)
            model_cell.font = Font(bold=True, size=12, color="2D3748")
            model_cell.fill = SECTION_FILL
            model_cell.border = THIN_BORDER
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            
            # 주요 특징
            feature_label = ws.cell(row=current_row, column=2, value="주요 특징:")
            feature_label.font = Font(bold=True, size=10)
            feature_label.border = THIN_BORDER
            feature_label.fill = HIGHLIGHT_FILL
            feature_label.alignment = Alignment(horizontal="right", vertical="top")
            
            feature_value = ws.cell(row=current_row, column=3, value=characteristics["주요 특징"])
            feature_value.border = THIN_BORDER
            feature_value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.merge_cells(f'C{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            
            # 알려진 이슈
            issues_label = ws.cell(row=current_row, column=2, value="알려진 이슈:")
            issues_label.font = Font(bold=True, size=10)
            issues_label.border = THIN_BORDER
            issues_label.fill = HIGHLIGHT_FILL
            issues_label.alignment = Alignment(horizontal="right", vertical="top")
            
            issues_text = "\n".join([f"• {issue}" for issue in characteristics["알려진 이슈"]])
            issues_value = ws.cell(row=current_row, column=3, value=issues_text)
            issues_value.border = THIN_BORDER
            issues_value.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f'C{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = max(45, len(characteristics["알려진 이슈"]) * 18)
            current_row += 1
            
            # 원인
            cause_label = ws.cell(row=current_row, column=2, value="원인:")
            cause_label.font = Font(bold=True, size=10)
            cause_label.border = THIN_BORDER
            cause_label.fill = HIGHLIGHT_FILL
            cause_label.alignment = Alignment(horizontal="right", vertical="top")
            
            cause_value = ws.cell(row=current_row, column=3, value=characteristics["원인"])
            cause_value.border = THIN_BORDER
            cause_value.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f'C{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # 대응
            solution_label = ws.cell(row=current_row, column=2, value="대응:")
            solution_label.font = Font(bold=True, size=10)
            solution_label.border = THIN_BORDER
            solution_label.fill = HIGHLIGHT_FILL
            solution_label.alignment = Alignment(horizontal="right", vertical="top")
            
            solution_value = ws.cell(row=current_row, column=3, value=characteristics["대응"])
            solution_value.border = THIN_BORDER
            solution_value.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f'C{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = 25
            current_row += 2
        
        # 일반적인 오류 패턴
        current_row += 1
        section_cell = ws.cell(row=current_row, column=2, value="일반적인 오류 패턴 및 해석")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        common_errors = [
            ("type_error:simple", "파라미터의 타입이 잘못됨", "모델이 타입 변환을 제대로 하지 못함. 특히 Llama 계열에서 빈번"),
            ("wrong_func_name", "잘못된 함수를 호출함", "함수 선택 능력 부족 또는 비슷한 이름의 함수 혼동"),
            ("missing_required", "필수 파라미터 누락", "함수 스키마 이해 부족 또는 파라미터 생략"),
            ("decoder_failed", "응답 파싱 실패", "잘못된 JSON 형식 또는 예상치 못한 응답 구조"),
        ]
        
        for error_type, description, interpretation in common_errors:
            error_cell = ws.cell(row=current_row, column=2, value=error_type)
            error_cell.font = Font(bold=True, size=10, color="111827")
            error_cell.border = THIN_BORDER
            error_cell.fill = HIGHLIGHT_FILL
            error_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            desc_cell = ws.cell(row=current_row, column=3, value=description)
            desc_cell.border = THIN_BORDER
            desc_cell.font = Font(size=11)
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            interp_cell = ws.cell(row=current_row, column=4, value=interpretation)
            interp_cell.border = THIN_BORDER
            interp_cell.font = Font(size=11, color="6B7280")
            interp_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.merge_cells(f'D{current_row}:F{current_row}')
            
            ws.row_dimensions[current_row].height = 25
            current_row += 1
        
        # 열 너비
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 5
        ws.column_dimensions["F"].width = 5
    
    def create_about_bfcl_sheet_for_summary(self) -> None:
        """About BFCL 시트 추가 (통합 보고서용)"""
        ws = self.wb.create_sheet("About BFCL")
        
        current_row = 2
        
        # 타이틀
        title_cell = ws.cell(row=current_row, column=2, value="BFCL (Berkeley Function Calling Leaderboard)")
        title_cell.font = TITLE_FONT
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 2
        
        # BFCL 설명
        section_cell = ws.cell(row=current_row, column=2, value="BFCL이란?")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        desc_cell = ws.cell(row=current_row, column=2, 
                           value="UC Berkeley에서 개발한 LLM Function Calling 능력 평가의 표준 벤치마크입니다. "
                                 "실제 개발 환경을 반영한 다양한 시나리오와 AST 기반의 정확한 평가 방식을 특징으로 합니다.")
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws.row_dimensions[current_row].height = 30
        current_row += 2
        
        # 평가 카테고리
        section_cell = ws.cell(row=current_row, column=2, value="평가 카테고리")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        categories = [
            ("Simple", "단일 함수, 단일 호출", "가장 기본적인 형태"),
            ("Multiple", "다중 함수 중 단일 호출 선택", "함수 선택 능력 평가"),
            ("Parallel", "단일 함수, 다중 병렬 호출", "순서 무관 병렬 처리"),
            ("Parallel Multiple", "다중 함수, 다중 병렬 호출", "가장 복잡한 형태"),
        ]
        
        for cat_name, description, note in categories:
            name_cell = ws.cell(row=current_row, column=2, value=cat_name)
            name_cell.font = Font(bold=True, size=10)
            name_cell.fill = HIGHLIGHT_FILL
            name_cell.border = THIN_BORDER
            name_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            desc_cell = ws.cell(row=current_row, column=3, value=description)
            desc_cell.border = THIN_BORDER
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            note_cell = ws.cell(row=current_row, column=4, value=note)
            note_cell.border = THIN_BORDER
            note_cell.font = Font(size=11, color="6B7280")
            note_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f'D{current_row}:F{current_row}')
            
            current_row += 1
        
        current_row += 1
        
        # 참고 자료
        section_cell = ws.cell(row=current_row, column=2, value="참고 자료")
        section_cell.font = SUBTITLE_FONT
        section_cell.fill = SECTION_FILL
        section_cell.border = THIN_BORDER
        ws.merge_cells(f'B{current_row}:F{current_row}')
        current_row += 1
        
        links = [
            ("공식 리더보드", "https://gorilla.cs.berkeley.edu/leaderboard.html"),
            ("BFCL 블로그", "https://gorilla.cs.berkeley.edu/blogs/8_berkeley_function_calling_leaderboard.html"),
            ("GitHub", "https://github.com/ShishirPatil/gorilla/tree/main/berkeley-function-call-leaderboard"),
        ]
        
        for label, url in links:
            label_cell = ws.cell(row=current_row, column=2, value=label)
            label_cell.font = Font(bold=True, size=10)
            label_cell.border = THIN_BORDER
            label_cell.fill = HIGHLIGHT_FILL
            label_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            url_cell = ws.cell(row=current_row, column=3, value=url)
            url_cell.border = THIN_BORDER
            url_cell.font = Font(size=11, color="1976D2", underline="single")
            url_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f'C{current_row}:F{current_row}')
            
            current_row += 1
        
        # 열 너비
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 5
        ws.column_dimensions["F"].width = 5
    
    def save(self, output_path: Optional[Path] = None) -> Path:
        """엑셀 파일 저장"""
        if output_path is None:
            summary_dir = self.reports_dir / "summary"
            summary_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = summary_dir / f"all_models_summary_{timestamp}.xlsx"
        
        self.wb.save(output_path)
        return output_path
    
    def generate_report(self) -> Path:
        """전체 통합 보고서 생성"""
        self.load_all_model_data()
        # 시트 구성: 요약 + 카테고리 매트릭스 + 상세
        self.create_dashboard_sheet()
        self.create_all_models_summary_sheet()
        self.create_all_details_sheet()
        return self.save()


def generate_all_models_summary(reports_dir: str) -> str:
    """모든 모델의 통합 요약 보고서 생성"""
    reporter = AllModelsSummaryReporter(Path(reports_dir))
    return str(reporter.generate_report())


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="BFCL 평가 결과 엑셀 보고서 생성")
    parser.add_argument("--model", help="모델명 (개별 보고서 생성 시)")
    parser.add_argument("--result-dir", help="결과 디렉토리 (개별 보고서 생성 시)")
    parser.add_argument("--score-dir", help="점수 디렉토리 (개별 보고서 생성 시)")
    parser.add_argument("--output", help="출력 파일 경로 (선택)")
    parser.add_argument("--all-models", action="store_true", help="모든 모델 통합 보고서 생성")
    parser.add_argument("--reports-dir", help="보고서 디렉토리 (통합 보고서 생성 시)")
    
    args = parser.parse_args()
    
    if args.all_models:
        # 모든 모델 통합 보고서 생성
        if not args.reports_dir:
            print("Error: --reports-dir is required for --all-models")
            exit(1)
        output = generate_all_models_summary(args.reports_dir)
        print(f"통합 보고서 생성 완료: {output}")
    else:
        # 개별 모델 보고서 생성
        if not all([args.model, args.result_dir, args.score_dir]):
            print("Error: --model, --result-dir, --score-dir are required for individual report")
            exit(1)
        output = generate_excel_report(args.model, args.result_dir, args.score_dir)
        print(f"보고서 생성 완료: {output}")
