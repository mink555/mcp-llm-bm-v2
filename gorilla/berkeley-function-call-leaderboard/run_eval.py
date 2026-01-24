#!/usr/bin/env python3
"""
BFCL 평가 자동화 스크립트

사용법:
    # 퀵 테스트 (각 카테고리 2개씩)
    python run_eval.py --quick
    
    # 특정 모델만 퀵 테스트
    python run_eval.py --quick --models openrouter/qwen3-14b-FC
    
    # 전체 테스트
    python run_eval.py --full
    
    # 특정 카테고리만
    python run_eval.py --quick --categories simple_python,multiple

결과물:
    reports/
    ├── {model_name}/
    │   └── {model_name}_eval_report.xlsx
    └── summary/
        └── all_models_summary.xlsx
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# 지원 모델 목록
SUPPORTED_MODELS = [
    "openrouter/llama-3.3-70b-instruct-FC",
    "openrouter/mistral-small-3.2-24b-instruct-FC",
    "openrouter/qwen3-32b-FC",
    "openrouter/qwen3-14b-FC",
    "openrouter/qwen3-next-80b-a3b-instruct-FC",
]

# 퀵 테스트용 ID (각 카테고리 2개씩)
QUICK_TEST_IDS = {
    "simple_python": ["simple_python_0", "simple_python_1"],
    "multiple": ["multiple_0", "multiple_1"],
    "parallel": ["parallel_0", "parallel_1"],
}

# 전체 카테고리
ALL_CATEGORIES = ["simple_python", "multiple", "parallel"]


def run_command(cmd: list, description: str = "") -> bool:
    """명령 실행"""
    if description:
        print(f"\n{'='*60}")
        print(f"📌 {description}")
        print(f"{'='*60}")
    
    print(f"$ {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=False)
    return result.returncode == 0


def clean_directories():
    """이전 결과 정리"""
    print("\n🧹 이전 결과 정리 중...")
    for dir_name in ["result", "score"]:
        dir_path = Path(dir_name)
        if dir_path.exists():
            shutil.rmtree(dir_path)
        dir_path.mkdir(parents=True, exist_ok=True)
    print("✅ 정리 완료")


def setup_quick_test():
    """퀵 테스트 설정 파일 생성"""
    test_ids_file = Path("test_case_ids_to_generate.json")
    with open(test_ids_file, "w") as f:
        json.dump(QUICK_TEST_IDS, f, indent=2)
    print(f"✅ 퀵 테스트 설정 완료: {test_ids_file}")


def generate_results(model: str, categories: list, is_quick: bool) -> bool:
    """모델 응답 생성"""
    cmd = [
        "python", "-m", "bfcl_eval", "generate",
        "--model", model,
        "--test-category", ",".join(categories),
        "--temperature", "0",
        "--num-threads", "1",
    ]
    
    if is_quick:
        cmd.append("--run-ids")
    
    return run_command(cmd, f"응답 생성: {model}")


def evaluate_results(models: list, categories: list) -> bool:
    """평가 실행"""
    cmd = [
        "python", "-m", "bfcl_eval", "evaluate",
        "--model", ",".join(models),
        "--test-category", ",".join(categories),
        "--partial-eval",
    ]
    
    return run_command(cmd, "평가 실행")


def generate_reports(models: list):
    """엑셀 보고서 생성"""
    from excel_reporter import BFCLExcelReporter
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    
    # 스타일 정의
    HEADER_FILL = PatternFill(start_color="D9E2EC", end_color="D9E2EC", fill_type="solid")
    HEADER_FONT = Font(bold=True, size=11)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    PASS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    print(f"\n{'='*60}")
    print("📊 엑셀 보고서 생성")
    print(f"{'='*60}")
    
    # 모델별 보고서 생성
    model_reports = []
    all_results = {}  # 전체 취합용
    
    for model in models:
        safe_name = model.replace("/", "_")
        result_dir = Path(f"result/{safe_name}")
        score_dir = Path(f"score/{safe_name}")
        
        if not result_dir.exists():
            print(f"⚠️  결과 없음: {model}")
            continue
        
        # 모델별 폴더 생성
        model_report_dir = reports_dir / safe_name
        model_report_dir.mkdir(exist_ok=True)
        
        # 보고서 생성
        reporter = BFCLExcelReporter(
            model_name=model,
            result_dir=result_dir,
            score_dir=score_dir,
        )
        reporter.load_data()
        reporter.create_evaluation_criteria_sheet()
        reporter.create_detail_sheet()
        reporter.create_summary_sheet()
        
        report_path = model_report_dir / f"{safe_name}_eval_report.xlsx"
        reporter.wb.save(report_path)
        
        print(f"✅ {model}: {report_path}")
        model_reports.append(report_path)
        
        # 취합용 데이터 수집
        all_results[model] = {
            "categories": reporter.categories_found,
            "detail_data": reporter.detail_data,
        }
    
    # 전체 취합 보고서 생성
    if len(all_results) > 1:
        summary_dir = reports_dir / "summary"
        summary_dir.mkdir(exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Model Comparison"
        
        # 헤더
        headers = ["Model", "Category", "Total", "Correct", "Incorrect", "Accuracy"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        
        row = 2
        for model, data in all_results.items():
            # 카테고리별 통계 계산
            category_stats = {}
            for entry in data["detail_data"]:
                cat = entry["category"]
                if cat not in category_stats:
                    category_stats[cat] = {"total": 0, "correct": 0}
                category_stats[cat]["total"] += 1
                if entry["result"] == "PASS":
                    category_stats[cat]["correct"] += 1
            
            for cat, stats in sorted(category_stats.items()):
                total = stats["total"]
                correct = stats["correct"]
                incorrect = total - correct
                accuracy = correct / total if total > 0 else 0
                
                ws.cell(row=row, column=1, value=model).border = THIN_BORDER
                ws.cell(row=row, column=2, value=cat).border = THIN_BORDER
                ws.cell(row=row, column=3, value=total).border = THIN_BORDER
                ws.cell(row=row, column=4, value=correct).border = THIN_BORDER
                ws.cell(row=row, column=5, value=incorrect).border = THIN_BORDER
                
                acc_cell = ws.cell(row=row, column=6, value=accuracy)
                acc_cell.number_format = "0.00%"
                acc_cell.border = THIN_BORDER
                if accuracy >= 0.8:
                    acc_cell.fill = PASS_FILL
                elif accuracy < 0.5:
                    acc_cell.fill = FAIL_FILL
                
                row += 1
        
        # 열 너비 조정
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12
        
        summary_path = summary_dir / "all_models_summary.xlsx"
        wb.save(summary_path)
        print(f"✅ 전체 취합: {summary_path}")
    
    return model_reports


def print_summary(models: list):
    """결과 요약 출력"""
    print(f"\n{'='*60}")
    print("📋 평가 결과 요약")
    print(f"{'='*60}")
    
    # score 파일에서 결과 읽기
    for model in models:
        safe_name = model.replace("/", "_")
        score_dir = Path(f"score/{safe_name}/non_live")
        
        if not score_dir.exists():
            continue
        
        print(f"\n🦍 {model}")
        
        for score_file in sorted(score_dir.glob("*_score.json")):
            with open(score_file) as f:
                first_line = f.readline()
                summary = json.loads(first_line)
            
            category = score_file.name.replace("BFCL_v4_", "").replace("_score.json", "")
            accuracy = summary.get("accuracy", 0) * 100
            correct = summary.get("correct_count", 0)
            total = summary.get("total_count", 0)
            
            status = "✅" if accuracy >= 80 else "⚠️" if accuracy >= 50 else "❌"
            print(f"   {status} {category}: {accuracy:.1f}% ({correct}/{total})")


def main():
    parser = argparse.ArgumentParser(description="BFCL 평가 자동화")
    parser.add_argument("--quick", action="store_true", help="퀵 테스트 (각 카테고리 2개씩)")
    parser.add_argument("--full", action="store_true", help="전체 테스트")
    parser.add_argument("--models", type=str, help="테스트할 모델 (쉼표 구분)")
    parser.add_argument("--categories", type=str, help="테스트할 카테고리 (쉼표 구분)")
    parser.add_argument("--skip-generate", action="store_true", help="생성 단계 건너뛰기")
    parser.add_argument("--skip-evaluate", action="store_true", help="평가 단계 건너뛰기")
    parser.add_argument("--report-only", action="store_true", help="보고서만 생성")
    
    args = parser.parse_args()
    
    # 기본값: 퀵 테스트
    if not args.quick and not args.full:
        args.quick = True
    
    # 모델 목록
    if args.models:
        models = [m.strip() for m in args.models.split(",")]
    else:
        models = SUPPORTED_MODELS
    
    # 카테고리 목록
    if args.categories:
        categories = [c.strip() for c in args.categories.split(",")]
    else:
        categories = ALL_CATEGORIES
    
    print(f"""
╔══════════════════════════════════════════════════════════════╗
║                    BFCL 평가 자동화                          ║
╠══════════════════════════════════════════════════════════════╣
║  모드: {'퀵 테스트' if args.quick else '전체 테스트'}                                          ║
║  모델: {len(models)}개                                                ║
║  카테고리: {', '.join(categories):<43} ║
╚══════════════════════════════════════════════════════════════╝
""")
    
    # 1. 정리
    if not args.report_only and not args.skip_generate:
        clean_directories()
        
        if args.quick:
            setup_quick_test()
    
    # 2. 생성
    if not args.report_only and not args.skip_generate:
        for model in models:
            if not generate_results(model, categories, args.quick):
                print(f"❌ 생성 실패: {model}")
    
    # 3. 평가
    if not args.report_only and not args.skip_evaluate:
        if not evaluate_results(models, categories):
            print("❌ 평가 실패")
    
    # 4. 보고서 생성
    generate_reports(models)
    
    # 5. 요약 출력
    print_summary(models)
    
    print(f"\n{'='*60}")
    print("🎉 완료!")
    print(f"{'='*60}")
    print(f"📁 보고서 위치: reports/")


if __name__ == "__main__":
    main()
